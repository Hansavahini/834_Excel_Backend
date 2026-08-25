"""
EDI endpoints.

The through-line of this module is that the backend, not the browser, decides
what is true. Three specific consequences, each of which was previously the
other way round:

  * A file is synced into the member tables only when validation passed. The
    old code set QUARANTINED and then called sync_uploaded_file() anyway, so an
    835 or a truncated 834 still wrote eligibility history.

  * Conversion checks the stored validation status itself rather than trusting
    that the UI disabled a button. The convert endpoint is reachable with any
    file id an authenticated user can guess at.

  * The workbook the user downloads is the workbook the server generated. There
    is no second conversion engine anywhere in the request path.

Ownership is checked on every object retrieval, and where the deployment has
clients configured the selected client narrows it further.
"""

from __future__ import annotations

import logging
import os
from datetime import datetime

from django.db import transaction
from django.http import FileResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from django.conf import settings

from conversion.models import ConversionHistory
from files.models import (
    VALIDATED_STATUSES,
    GeneratedFile,
    ProcessingStatus,
    UploadedFile,
    canonical_status,
)
from files.models import sha256_of
from mapping.models import SegmentElement
from members.models import Dependant, Subscriber
from users.tenancy import resolve_client, scope_to_client

from edi.services.excel_generator import generate_excel
from edi.services.file_service import UnsafePathError, get_file_path
from edi.services.ingest import sync_uploaded_file
from edi.services.loop_extractor import StreamingParsedFile
from edi.services.mapping_store import (
    get_mappings,
    get_template,
    headers_for,
    lock_template,
    save_mapping,
)
from edi.services.parser import EDI834Parser, EDIParseError, envelope_facts
from edi.services.row_builder import column_kinds, iter_excel_rows
from edi.services.validator import validate_834

from .serializers import ConvertRequestSerializer, EDIFileUploadSerializer, MappingSerializer

logger = logging.getLogger("edi.api")

# The only statuses from which a file may be converted. VALIDATED is the
# canonical one; PARSED is its legacy spelling and CONVERTED is here because
# re-running a conversion after a mapping change is a normal thing to do and
# refusing it would mean re-uploading the file to change one column.
CONVERTIBLE_STATUSES = VALIDATED_STATUSES

MAX_PREVIEW_ROWS = 500

# Streaming chunk for the raw-source download. Large enough that a 200 MB file
# is not a million syscalls, small enough that it is never resident.
DOWNLOAD_CHUNK = 64 * 1024

# How many characters of an 834 the preview endpoint returns. The preview is a
# convenience for the on-screen viewer; the download endpoint is the file.
PREVIEW_CHAR_LIMIT = 512 * 1024

DISPLAY_DATE_FORMAT = getattr(settings, "DISPLAY_DATE_FORMAT", "%m-%d-%Y")


def display_date(value):
    """
    MM-DD-YYYY, or empty.

    Part 4 asked for one date format across the workbook, the API and the
    screen. Dates still travel as ISO in their own fields so anything that
    sorts or filters keeps working; these are the strings the UI renders, which
    is what stops each component inventing its own toLocaleDateString call and
    each one getting a slightly different answer.
    """
    if not value:
        return ""
    try:
        return value.strftime(DISPLAY_DATE_FORMAT)
    except AttributeError:
        return str(value)


def _parse_x12_date(value: str):
    value = (value or "").strip()
    if len(value) == 6:  # GS04 in some 004010 files is YYMMDD
        value = "20" + value
    if len(value) != 8 or not value.isdigit():
        return None
    try:
        return datetime.strptime(value, "%Y%m%d").date()
    except ValueError:
        return None


def _mapping_snapshot(rules):
    """
    The rules that ran, flattened to plain JSON.

    Accepts either MappingDetail rows or the dicts the API takes, because both
    reach the converter and both have to be recordable. Stored on the history
    row so "which mapping produced this workbook" has an answer that does not
    depend on the template still existing in its original form.
    """
    snapshot = []
    for rule in rules:
        if isinstance(rule, dict):
            snapshot.append(
                {
                    "excel_column": rule.get("excel_column", ""),
                    "segment": rule.get("segment", ""),
                    "element": rule.get("element", ""),
                    "qualifier_element": rule.get("qualifier_element", "") or "",
                    "qualifier_value": rule.get("qualifier_value", "") or "",
                    "occurrence": rule.get("occurrence") or 1,
                    "applies_to": rule.get("applies_to") or "BOTH",
                    "transform": rule.get("transform") or "NONE",
                }
            )
        else:
            snapshot.append(
                {
                    "excel_column": rule.excel_column,
                    "segment": rule.segment,
                    "element": rule.element,
                    "qualifier_element": rule.qualifier_element or "",
                    "qualifier_value": rule.qualifier_value or "",
                    "occurrence": rule.occurrence or 1,
                    "applies_to": rule.applies_to,
                    "transform": rule.transform,
                }
            )
    return snapshot


def owned_uploads(request, client):
    """Every uploaded-file query in this module starts here."""
    return scope_to_client(
        UploadedFile.objects.filter(owner=request.user), client
    )


def owned_generated(request, client):
    return scope_to_client(
        GeneratedFile.objects.filter(owner=request.user), client
    )


def _upload_payload(record):
    """
    One uploaded file, complete enough that a browser refresh restores the screen.

    Issue 1 in full. The row already held the parse result; what it did not hold
    was the validation outcome, the conversion outcome or a route back to the
    workbook, so those three lived in React and died with it. Everything the
    conversion screen renders now comes from here, which means the answer to
    "what happened to this file" is the same whether you ask it a second after
    the upload or a week later from another machine.
    """
    latest = record.generated_files.order_by("-generated_at").first()
    status = canonical_status(record.processing_status)
    return {
        "id": record.id,
        "uploaded_file_id": record.id,
        "fileName": record.original_filename,
        "original_filename": record.original_filename,
        "status": status,
        "is_valid": record.processing_status in VALIDATED_STATUSES,
        "is_converted": status == ProcessingStatus.CONVERTED,
        "records": record.member_loop_count or 0,
        "member_loop_count": record.member_loop_count,
        "segment_count": record.segment_count,
        "is_full_file": record.is_full_file,
        "file_date": record.file_date,
        "file_date_display": display_date(record.file_date),
        "file_size_bytes": record.file_size_bytes,
        "file_path": record.stored_file.name,
        "uploaded_at": record.uploaded_at,
        "uploaded_at_display": display_date(record.uploaded_at),
        "validated_at": record.validated_at,
        "validated_at_display": display_date(record.validated_at),
        "converted_at": record.converted_at,
        "converted_at_display": display_date(record.converted_at),
        "error_message": record.error_message,
        "conversion_error": record.conversion_error,
        "errors": record.validation_errors or [],
        "warnings": record.validation_warnings or [],
        "sponsor_name": record.sponsor_name,
        "generated_file_id": latest.id if latest else None,
        "generated_filename": latest.generated_filename if latest else None,
        "generated_row_count": latest.row_count if latest else None,
        "download_url": (
            "/api/edi/download/{pk}/".format(pk=latest.id) if latest else None
        ),
        "preview_url": (
            "/api/edi/download/{pk}/preview/".format(pk=latest.id) if latest else None
        ),
        # Part 8: preview and download are different endpoints because they are
        # different jobs. One returns a readable excerpt, the other returns the
        # bytes that arrived.
        "source_url": "/api/edi/files/{pk}/preview/".format(pk=record.id),
        "source_download_url": "/api/edi/files/{pk}/download/".format(pk=record.id),
    }


class HealthCheckView(APIView):
    permission_classes = []  # a health probe cannot authenticate

    def get(self, request):
        return Response({"status": "healthy", "service": "834 EDI Converter"})


class EDIUploadView(APIView):
    """Store the file, checksum it, validate it, and record what arrived."""

    def post(self, request):
        serializer = EDIFileUploadSerializer(data=request.data)

        if not serializer.is_valid():
            logger.error("Upload validation failed: %s", serializer.errors)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        client = resolve_client(request)
        upload = serializer.validated_data["file"]

        # ---------------------------------------------------------------
        # Issue 9: MAX_834_UPLOAD_BYTES was a setting nothing read.
        #
        # Checked here rather than only in the serializer because a chunked
        # upload can report one size and deliver another, and because the
        # message wants to name the actual limit — "file too large" with no
        # number is the kind of error that generates a support ticket instead
        # of resolving one.
        # ---------------------------------------------------------------
        limit = getattr(settings, "MAX_834_UPLOAD_BYTES", 200 * 1024 * 1024)
        size = getattr(upload, "size", None) or 0
        if size > limit:
            return Response(
                {
                    "detail": (
                        "{name} is {actual:.1f} MB. The largest 834 this portal accepts "
                        "is {limit:.0f} MB. Split the interchange or ask an administrator "
                        "to raise MAX_834_UPLOAD_BYTES."
                    ).format(
                        name=upload.name,
                        actual=size / (1024 * 1024),
                        limit=limit / (1024 * 1024),
                    ),
                    "file_size_bytes": size,
                    "max_upload_bytes": limit,
                },
                status=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            )

        checksum = sha256_of(upload)

        existing = (
            UploadedFile.objects.filter(owner=request.user, content_sha256=checksum)
            .filter(client=client)
            .first()
        )

        if existing and existing.processing_status not in (
            ProcessingStatus.FAILED,
            ProcessingStatus.PENDING,
            ProcessingStatus.UPLOADED,
            ProcessingStatus.PARSING,
            ProcessingStatus.VALIDATING,
        ):
            # A file that was processed to a conclusion, valid or quarantined.
            # Re-uploading it is a no-op, which is the point of the checksum.
            return Response(
                {
                    "message": "This file has already been uploaded.",
                    "uploaded_file_id": existing.id,
                    "file_path": existing.stored_file.name,
                    "status": canonical_status(existing.processing_status),
                    "is_valid": existing.processing_status in VALIDATED_STATUSES,
                    "duplicate": True,
                },
                status=status.HTTP_200_OK,
            )

        if existing:
            # Issue 19: the first attempt died. Duplicate protection was
            # rejecting the retry, which left the user with a permanently
            # unusable file and no way to clear it — the checksum matched, so
            # every subsequent upload bounced off the same broken record.
            # Reuse the row so the audit trail keeps one history for one file.
            record = existing
            record.stored_file.save(upload.name, upload, save=False)
            record.original_filename = upload.name[:255]
            record.file_size_bytes = upload.size
            record.error_message = ""
            record.conversion_error = ""
            record.validation_errors = []
            record.validation_warnings = []
            record.validated_at = None
            record.processing_status = ProcessingStatus.UPLOADED
            record.processing_started_at = None
            record.processing_finished_at = None
            record.save()
            retried = True
        else:
            record = UploadedFile(
                owner=request.user,
                client=client,
                original_filename=upload.name[:255],
                file_size_bytes=upload.size,
                content_sha256=checksum,
                processing_status=ProcessingStatus.UPLOADED,
            )
            record.stored_file.save(upload.name, upload, save=False)
            record.save()
            retried = False

        return self._process(request, record, client, retried)

    def _process(self, request, record, client, retried):
        try:
            record.processing_status = ProcessingStatus.VALIDATING
            record.processing_started_at = timezone.now()
            record.save(update_fields=["processing_status", "processing_started_at"])

            parser = EDI834Parser(record.stored_file.path)

            header = []
            header_done = False

            def capture_header(stream):
                nonlocal header_done
                for segment in stream:
                    if not header_done:
                        if segment.name == "INS":
                            header_done = True
                        else:
                            header.append(segment)
                    yield segment

            result = validate_834(capture_header(parser.iter_segments()))

            if result is None:
                raise RuntimeError(
                    "validate_834() returned None instead of a validation result."
                )

            facts = envelope_facts(header)

            for field_name in (
                "interchange_control_number",
                "group_control_number",
                "transaction_set_control_number",
                "sender_id",
                "receiver_id",
                "sponsor_name",
            ):
                setattr(record, field_name, facts[field_name][:60])

            record.file_date = _parse_x12_date(facts["file_date"])
            record.segment_count = result.segment_count
            record.member_loop_count = result.member_loop_count
            record.is_full_file = result.is_full_file

            record.processing_status = (
                ProcessingStatus.VALIDATED
                if result.is_valid
                else ProcessingStatus.QUARANTINED
            )
            # Persisted, not just returned. This is the half of Issue 1 that
            # made a refresh destructive: the answer existed for the duration of
            # one HTTP response and was then discarded.
            record.validation_errors = list(result.errors)[:200]
            record.validation_warnings = list(result.warnings)[:200]
            record.validated_at = timezone.now()
            record.error_message = "\n".join(result.errors)[:4000]
            record.processing_finished_at = timezone.now()
            record.save()

            # ---------------------------------------------------------
            # DATABASE SYNC ENGINE — valid files only.
            #
            # The old code ran this unconditionally, right after setting the
            # status to QUARANTINED. So an 835, a truncated interchange or a
            # file with mismatched control numbers still wrote Member and
            # MemberDailyStatus rows, and those rows then answered eligibility
            # questions as if the file had been trustworthy. Quarantine that
            # does not actually quarantine anything is worse than none, because
            # it reads as a control that is working.
            # ---------------------------------------------------------
            sync_summary = {"synced": 0, "failed": 0, "loops": 0, "errors": []}
            if result.is_valid:
                try:
                    sync_summary = sync_uploaded_file(record, request.user, client)
                except Exception as exc:  # noqa: BLE001
                    logger.exception("Member sync failed for upload %s", record.id)
                    sync_summary["errors"] = [
                        "{k}: {e}".format(k=type(exc).__name__, e=exc)
                    ]
            else:
                logger.warning(
                    "Upload %s failed validation; not synced to the member tables.",
                    record.id,
                )

            return Response(
                {
                    "message": (
                        "File uploaded and parsed successfully."
                        if result.is_valid
                        else "File was rejected by 834 validation and has been quarantined."
                    ),
                    "uploaded_file_id": record.id,
                    "status": canonical_status(record.processing_status),
                    "member_loop_count": record.member_loop_count,
                    "segment_count": record.segment_count,
                    "is_full_file": record.is_full_file,
                    "file_path": record.stored_file.name,
                    "file_date": record.file_date,
                    "file_date_display": display_date(record.file_date),
                    "duplicate": False,
                    "retried": retried,
                    "is_valid": result.is_valid,
                    "errors": result.errors[:25],
                    "warnings": result.warnings[:25],
                    "transaction_count": result.transaction_count,
                    "members_synced": sync_summary.get("synced", 0),
                    "members_failed": sync_summary.get("failed", 0),
                    "sync_errors": sync_summary.get("errors", [])[:10],
                },
                status=status.HTTP_201_CREATED,
            )

        except (EDIParseError, OSError) as exc:
            record.processing_status = ProcessingStatus.FAILED
            record.error_message = str(exc)[:4000]
            record.validation_errors = [str(exc)[:500]]
            record.validated_at = timezone.now()
            record.processing_finished_at = timezone.now()
            record.save(
                update_fields=[
                    "processing_status",
                    "error_message",
                    "validation_errors",
                    "validated_at",
                    "processing_finished_at",
                ]
            )
            logger.warning("upload %s failed to parse: %s", record.id, exc)
            return Response(
                {
                    "message": "File stored but could not be parsed.",
                    "uploaded_file_id": record.id,
                    "status": canonical_status(record.processing_status),
                    "is_valid": False,
                    "error": str(exc),
                    "retryable": True,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        except Exception as exc:  # noqa: BLE001
            record.processing_status = ProcessingStatus.FAILED
            record.error_message = "{t}: {e}".format(t=type(exc).__name__, e=exc)[:4000]
            record.validation_errors = [record.error_message[:500]]
            record.validated_at = timezone.now()
            record.processing_finished_at = timezone.now()
            record.save(
                update_fields=[
                    "processing_status",
                    "error_message",
                    "validation_errors",
                    "validated_at",
                    "processing_finished_at",
                ]
            )
            logger.exception("Unexpected error while processing upload %s", record.id)
            return Response(
                {
                    "message": "An unexpected error occurred while processing the file.",
                    "uploaded_file_id": record.id,
                    "status": canonical_status(record.processing_status),
                    "is_valid": False,
                    "retryable": True,
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class UploadListView(APIView):
    """
    The current user's uploads.

    The conversion screen kept its file list in React state alone, so a browser
    refresh emptied a table whose rows were sitting in the database the whole
    time. This is what it reloads from.
    """

    def get(self, request):
        client = resolve_client(request)
        records = (
            owned_uploads(request, client)
            .prefetch_related("generated_files")
            .order_by("-uploaded_at")[:200]
        )
        return Response([_upload_payload(record) for record in records])


class UploadSourceView(APIView):
    """
    The raw 834 as it was stored, for the EDI viewer.

    Owner and client checked here, because the viewer used to fall back to a
    hard-coded sample interchange when it had no content in hand. A real file
    name above somebody else's demo member data is a worse outcome than an
    error message.
    """

    def get(self, request, pk):
        client = resolve_client(request)
        record = owned_uploads(request, client).filter(pk=pk).first()
        if not record:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        try:
            with open(
                record.stored_file.path, "r", encoding="utf-8-sig", errors="replace"
            ) as handle:
                content = handle.read(4 * 1024 * 1024)
        except OSError as exc:
            logger.warning("Could not read stored source for upload %s: %s", pk, exc)
            return Response(
                {"detail": "The stored source file could not be read."},
                status=status.HTTP_404_NOT_FOUND,
            )

        return Response(
            {
                "uploaded_file_id": record.id,
                "file_name": record.original_filename,
                "status": record.processing_status,
                "truncated": len(content) >= 4 * 1024 * 1024,
                "content": content,
            }
        )


class ValidateView(APIView):
    """Structural validation on its own, so a user can check a file before converting."""

    def post(self, request):
        client = resolve_client(request)
        record = None

        if request.data.get("uploaded_file_id"):
            record = (
                owned_uploads(request, client)
                .filter(pk=request.data["uploaded_file_id"])
                .first()
            )
            if not record:
                return Response({"detail": "Unknown uploaded_file_id."}, status=404)
            path = record.stored_file.path
        else:
            try:
                path = get_file_path(request.data.get("file_path", ""))
            except UnsafePathError as exc:
                return Response(
                    {"file_path": str(exc)}, status=status.HTTP_400_BAD_REQUEST
                )

        if record:
            record.processing_status = ProcessingStatus.VALIDATING
            record.save(update_fields=["processing_status"])

        try:
            result = validate_834(EDI834Parser(path).iter_segments())
        except EDIParseError as exc:
            if record:
                record.processing_status = ProcessingStatus.QUARANTINED
                record.error_message = str(exc)[:4000]
                record.validation_errors = [str(exc)[:500]]
                record.validated_at = timezone.now()
                record.save(
                    update_fields=[
                        "processing_status",
                        "error_message",
                        "validation_errors",
                        "validated_at",
                    ]
                )
            return Response(
                {"is_valid": False, "errors": [str(exc)]}, status=status.HTTP_200_OK
            )

        payload = result.as_dict()

        if record:
            # Keep the stored status honest, and keep the result. A file that
            # fails validation here must not still be sitting at VALIDATED,
            # because that is what the convert endpoint checks — and the errors
            # must outlive the response, because that is what the screen shows
            # after a refresh.
            already_converted = (
                record.processing_status == ProcessingStatus.CONVERTED
                and result.is_valid
            )
            record.processing_status = (
                record.processing_status
                if already_converted
                else (
                    ProcessingStatus.VALIDATED
                    if result.is_valid
                    else ProcessingStatus.QUARANTINED
                )
            )
            record.validation_errors = list(result.errors)[:200]
            record.validation_warnings = list(result.warnings)[:200]
            record.validated_at = timezone.now()
            record.error_message = "\n".join(result.errors)[:4000]
            record.save(
                update_fields=[
                    "processing_status",
                    "error_message",
                    "validation_errors",
                    "validation_warnings",
                    "validated_at",
                ]
            )
            payload["uploaded_file_id"] = record.id
            payload["status"] = canonical_status(record.processing_status)
            payload["validated_at"] = record.validated_at
            payload["validated_at_display"] = display_date(record.validated_at)

        return Response(payload)


class MappingCreateView(APIView):
    """Save a column rule to the user's mapping template, and list what is saved."""

    def get(self, request):
        client = resolve_client(request)
        template_id = request.query_params.get("template_id")
        details = get_mappings(request.user, template_id, client=client)
        template = get_template(request.user, template_id, client=client)
        return Response(
            {
                "template": template.mapping_name if template else None,
                "template_id": template.id if template else None,
                "version": template.version if template else None,
                "locked": bool(template and template.is_locked),
                "columns": [
                    {
                        "excel_column": detail.excel_column,
                        "column_order": detail.column_order,
                        "segment": detail.segment,
                        "element": detail.element,
                        "qualifier_element": detail.qualifier_element,
                        "qualifier_value": detail.qualifier_value,
                        "occurrence": detail.occurrence,
                        "applies_to": detail.applies_to,
                        "transform": detail.transform,
                    }
                    for detail in details
                ],
            }
        )

    def post(self, request):
        client = resolve_client(request)
        many = isinstance(request.data, list)
        serializer = MappingSerializer(data=request.data, many=many)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        payload = serializer.validated_data if many else [serializer.validated_data]
        saved = []
        with transaction.atomic():
            for rule in payload:
                saved.append(
                    save_mapping(
                        rule,
                        owner=request.user,
                        template_name=rule.get("template_name", "Default"),
                        client=client,
                    )
                )

        return Response(
            {
                "message": "Mapping saved",
                "template_id": saved[0]["template_id"] if saved else None,
                "version": saved[0]["version"] if saved else None,
                "mappings": saved,
            },
            status=status.HTTP_201_CREATED,
        )


class Convert834View(APIView):
    """Run the full pipeline and record the result."""

    def post(self, request):
        serializer = ConvertRequestSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        data = serializer.validated_data
        client = resolve_client(request)

        # ---------------------------------------------------------------
        # Resolve the file through the owner-scoped queryset, always. The
        # endpoint is reachable with any id, and the front end blocking a
        # button is not an authorisation control.
        # ---------------------------------------------------------------
        if data.get("uploaded_file_id"):
            record = (
                owned_uploads(request, client)
                .filter(pk=data["uploaded_file_id"])
                .first()
            )
            if not record:
                return Response(
                    {"uploaded_file_id": "Unknown file."},
                    status=status.HTTP_404_NOT_FOUND,
                )
        else:
            try:
                get_file_path(data["file_path"])
            except UnsafePathError as exc:
                return Response(
                    {"file_path": str(exc)}, status=status.HTTP_400_BAD_REQUEST
                )
            record = (
                owned_uploads(request, client)
                .filter(stored_file=data["file_path"])
                .first()
            )
            if not record:
                return Response(
                    {
                        "file_path": (
                            "That path does not correspond to a file you uploaded. "
                            "Supply uploaded_file_id instead."
                        )
                    },
                    status=status.HTTP_404_NOT_FOUND,
                )

        # Issue 10: the backend enforces the validation gate itself.
        if record.processing_status not in CONVERTIBLE_STATUSES:
            return Response(
                {
                    "detail": (
                        "This file is in status {s} and cannot be converted. Only a file "
                        "that passed 834 validation may be converted."
                    ).format(s=record.processing_status),
                    "uploaded_file_id": record.id,
                    "status": record.processing_status,
                    "errors": (record.error_message or "").splitlines()[:25],
                },
                status=status.HTTP_422_UNPROCESSABLE_ENTITY,
            )

        source_path = record.stored_file.path

        # ---------------------------------------------------------------
        # Issue 6.2: a conversion must be able to say which rules produced it.
        #
        # The old behaviour: the browser posted ad-hoc rules, the run used
        # them, and ConversionHistory recorded mapping_template=None. The
        # saved templates in the database were ignored, and the audit trail
        # could say a workbook had been produced but not how — which is the
        # one question an audit trail exists to answer.
        #
        # Inline rules are now persisted to the user's template first, so the
        # run has a template id and a version like any other. The exact rules
        # are frozen onto the history row as well, because a template id only
        # resolves while the template still exists and a snapshot always does.
        # ---------------------------------------------------------------
        template = None
        mapping_source = "TEMPLATE"

        if data.get("mappings"):
            rules = data["mappings"]
            headers = data["headers"]
            mapping_source = "INLINE"
            try:
                with transaction.atomic():
                    for index, rule in enumerate(rules):
                        saved = save_mapping(
                            rule,
                            owner=request.user,
                            template_name=rule.get("template_name", "Default"),
                            client=client,
                        )
                    if rules:
                        template = get_template(request.user, saved["template_id"], client=client)
            except Exception:  # noqa: BLE001
                # Provenance is worth having but not worth failing a conversion
                # over. The snapshot below still records exactly what ran.
                logger.exception(
                    "Could not persist inline mapping rules for upload %s", record.id
                )
        else:
            template = get_template(
                request.user, data.get("mapping_template_id"), client=client
            )
            rules = get_mappings(
                request.user, data.get("mapping_template_id"), client=client
            )
            if not rules:
                return Response(
                    {"detail": "No mapping rules supplied and no saved template found."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            headers = data.get("headers") or headers_for(rules)

        snapshot = _mapping_snapshot(rules)
        kinds = column_kinds(rules)

        record.processing_status = ProcessingStatus.CONVERTING
        record.conversion_error = ""
        record.save(update_fields=["processing_status", "conversion_error"])

        history = ConversionHistory.objects.create(
            owner=request.user,
            client=client,
            uploaded_file=record,
            mapping_template=template,
            mapping_version=template.version if template else None,
            mapping_snapshot=snapshot,
            mapping_source=mapping_source,
            status=ConversionHistory.Status.RUNNING,
            started_at=timezone.now(),
        )

        warnings = []

        def rows_with_warnings(stream):
            """Drain warnings as rows go past so nothing is held for a second pass."""
            for row in stream:
                warnings.extend(row.pop("__warnings__", []))
                yield row

        try:
            parser = EDI834Parser(source_path)
            parsed = StreamingParsedFile(parser.iter_segments())

            rows = rows_with_warnings(
                iter_excel_rows(parsed, rules, header_segments=parsed.header)
            )

            workbook = generate_excel(
                headers,
                rows,
                owner_id=request.user.id,
                source_name=record.original_filename,
                # Part 4 and Part 14: the mapping decides which columns are
                # real dates and which must stay text, so the workbook carries
                # 08-25-2026 in a date cell and 001234567 with its leading zero.
                column_kinds=kinds,
            )
        except (EDIParseError, ValueError, OSError) as exc:
            history.status = ConversionHistory.Status.FAILED
            history.error_message = str(exc)[:4000]
            history.finished_at = timezone.now()
            history.save()
            # Back to VALIDATED, not stuck at CONVERTING. A status nothing can
            # leave is how a file becomes permanently unusable.
            record.processing_status = ProcessingStatus.VALIDATED
            record.conversion_error = str(exc)[:4000]
            record.save(update_fields=["processing_status", "conversion_error"])
            logger.exception("conversion failed for %s", source_path)
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        generated = GeneratedFile.objects.create(
            owner=request.user,
            client=client,
            uploaded_file=record,
            generated_filename=workbook.filename,
            stored_file=workbook.relative_path,
            row_count=workbook.row_count,
            file_size_bytes=workbook.size_bytes,
        )

        subscribers = parsed.subscriber_count
        history.generated_file = generated
        history.status = (
            ConversionHistory.Status.PARTIAL if warnings else ConversionHistory.Status.SUCCESS
        )
        history.members_processed = subscribers
        history.dependents_processed = parsed.dependent_count
        history.rows_written = workbook.row_count
        history.warning_count = len(warnings)
        history.warnings = warnings[:500]
        history.finished_at = timezone.now()
        history.save()

        # Issue 16: the version that produced this workbook is now frozen. A
        # later edit clones to the next version rather than rewriting what this
        # history row points at.
        lock_template(template)

        # Issue 1: CONVERTED is a stored fact now. It used to be a React
        # variable, which is why refreshing the page took the download link
        # with it even though the workbook was on disk and its GeneratedFile
        # row was in the database.
        record.processing_status = ProcessingStatus.CONVERTED
        record.converted_at = timezone.now()
        record.conversion_error = ""
        record.save(
            update_fields=["processing_status", "converted_at", "conversion_error"]
        )

        return Response(
            {
                "message": "834 converted successfully",
                "conversion_id": history.id,
                "uploaded_file_id": record.id,
                "status": canonical_status(record.processing_status),
                "converted_at": record.converted_at,
                "converted_at_display": display_date(record.converted_at),
                "generated_file_id": generated.id,
                "file": workbook.filename,
                "download_url": "/api/edi/download/{pk}/".format(pk=generated.id),
                "preview_url": "/api/edi/download/{pk}/preview/".format(pk=generated.id),
                "headers": list(headers),
                "rows_generated": workbook.row_count,
                "subscribers": subscribers,
                "dependents": parsed.dependent_count,
                "mapping_template_id": template.id if template else None,
                "mapping_version": template.version if template else None,
                "mapping_source": mapping_source,
                "mapping_columns": len(snapshot),
                "warning_count": len(warnings),
                "warnings": warnings[:25],
            }
        )


class DownloadView(APIView):
    """
    Retrieve a generated workbook.

    MEDIA_ROOT is not served by the URL conf, deliberately, because these files
    contain PHI and must not be reachable by guessing a path. Ownership is
    checked here and the download is counted for the audit trail.
    """

    def get(self, request, pk):
        client = resolve_client(request)
        generated = owned_generated(request, client).filter(pk=pk).first()
        if not generated:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        GeneratedFile.objects.filter(pk=pk).update(
            downloaded_count=generated.downloaded_count + 1,
            last_downloaded_at=timezone.now(),
        )
        return FileResponse(
            generated.stored_file.open("rb"),
            as_attachment=True,
            filename=generated.generated_filename,
        )


class GeneratedFilePreviewView(APIView):
    """
    Read the generated workbook back and return its cells as JSON.

    This exists so the Excel preview shows the file that was actually produced.
    The browser previously re-parsed the raw 834 with a second, qualifier-blind
    converter, which meant the preview and the download could disagree — and
    when they disagreed the preview was the one the user believed.
    """

    def get(self, request, pk):
        from openpyxl import load_workbook

        client = resolve_client(request)
        generated = owned_generated(request, client).filter(pk=pk).first()
        if not generated:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        try:
            limit = min(int(request.query_params.get("limit", 100)), MAX_PREVIEW_ROWS)
        except (TypeError, ValueError):
            limit = 100

        try:
            workbook = load_workbook(
                generated.stored_file.path, read_only=True, data_only=True
            )
        except (OSError, KeyError, ValueError) as exc:
            logger.warning("Could not read generated workbook %s: %s", pk, exc)
            return Response(
                {"detail": "The generated workbook could not be read."},
                status=status.HTTP_404_NOT_FOUND,
            )

        try:
            sheet = workbook.worksheets[0]
            iterator = sheet.iter_rows(values_only=True)
            headers = [
                "" if value is None else str(value) for value in next(iterator, ()) or ()
            ]
            rows = []
            for values in iterator:
                if len(rows) >= limit:
                    break
                rows.append(
                    {
                        header: ("" if value is None else str(value))
                        for header, value in zip(headers, values)
                    }
                )
        finally:
            workbook.close()

        return Response(
            {
                "generated_file_id": generated.id,
                "file_name": generated.generated_filename,
                "source_file": generated.uploaded_file.original_filename,
                "headers": headers,
                "rows": rows,
                "row_count": generated.row_count,
                "returned": len(rows),
                "truncated": bool(generated.row_count and generated.row_count > len(rows)),
                "download_url": "/api/edi/download/{pk}/".format(pk=generated.id),
            }
        )


class ConversionHistoryView(APIView):
    def get(self, request):
        client = resolve_client(request)
        history = scope_to_client(
            ConversionHistory.objects.filter(owner=request.user), client
        ).select_related("uploaded_file", "generated_file", "mapping_template")[:100]

        return Response(
            [
                {
                    "id": item.id,
                    "uploaded_file_id": item.uploaded_file_id,
                    "source_file": item.uploaded_file.original_filename,
                    "original_file": item.uploaded_file.original_filename,
                    "generated_file": (
                        item.generated_file.generated_filename
                        if item.generated_file_id
                        else None
                    ),
                    "generated_file_id": item.generated_file_id,
                    "file_date": item.uploaded_file.file_date,
                    "file_date_display": display_date(item.uploaded_file.file_date),
                    "created_at_display": display_date(item.created_at),
                    "mapping_template": (
                        item.mapping_template.mapping_name
                        if item.mapping_template
                        else None
                    ),
                    "mapping_version": item.mapping_version,
                    "mapping_source": item.mapping_source,
                    "mapping_columns": len(item.mapping_snapshot or []),
                    "status": item.status,
                    "rows_written": item.rows_written,
                    "members": item.members_processed,
                    "dependents": item.dependents_processed,
                    "warning_count": item.warning_count,
                    "error_message": item.error_message,
                    "created_at": item.created_at,
                    "download_url": (
                        "/api/edi/download/{pk}/".format(pk=item.generated_file_id)
                        if item.generated_file_id
                        else None
                    ),
                    "preview_url": (
                        "/api/edi/download/{pk}/preview/".format(pk=item.generated_file_id)
                        if item.generated_file_id
                        else None
                    ),
                    "source_url": "/api/edi/uploads/{pk}/content/".format(
                        pk=item.uploaded_file_id
                    ),
                }
                for item in history
            ]
        )


# ===========================================================================
# Part 8 — the raw 834: preview and download are separate endpoints.
#
# They were one endpoint that read four megabytes into a string and returned it
# as a JSON field. That is a reasonable preview and a broken download: anything
# larger was silently truncated, and the truncation was reported in a "truncated"
# flag the download path never looked at. A client who downloaded a 60 MB
# interchange received the first four megabytes of it, in a file with the right
# name, with no error anywhere. Splitting them makes the contract explicit —
# preview returns an excerpt and says so; download returns the file.
# ===========================================================================


class EDIFilePreviewView(APIView):
    """
    A readable excerpt of the stored 834, for the on-screen viewer.

    Truncation is expected here and is reported honestly, with the byte offsets
    so a caller can page through a large file rather than guessing.
    """

    def get(self, request, pk):
        client = resolve_client(request)
        record = owned_uploads(request, client).filter(pk=pk).first()
        if not record:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        try:
            limit = min(
                int(request.query_params.get("limit", PREVIEW_CHAR_LIMIT)),
                PREVIEW_CHAR_LIMIT,
            )
        except (TypeError, ValueError):
            limit = PREVIEW_CHAR_LIMIT
        try:
            offset = max(0, int(request.query_params.get("offset", 0)))
        except (TypeError, ValueError):
            offset = 0

        try:
            with open(
                record.stored_file.path, "r", encoding="utf-8-sig", errors="replace"
            ) as handle:
                if offset:
                    handle.seek(offset)
                content = handle.read(limit)
        except OSError as exc:
            logger.warning("Could not read stored source for upload %s: %s", pk, exc)
            return Response(
                {"detail": "The stored source file could not be read."},
                status=status.HTTP_404_NOT_FOUND,
            )

        total = record.file_size_bytes or 0
        returned = len(content)
        return Response(
            {
                "uploaded_file_id": record.id,
                "file_name": record.original_filename,
                "status": canonical_status(record.processing_status),
                "file_size_bytes": total,
                "offset": offset,
                "returned_chars": returned,
                "truncated": bool(total and offset + returned < total),
                "download_url": "/api/edi/files/{pk}/download/".format(pk=record.id),
                "content": content,
            }
        )


class EDIFileDownloadView(APIView):
    """
    The stored 834, whole, streamed.

    FileResponse with an open handle rather than read() into a Response: the
    file is never resident, which is the only way a 200 MB interchange is
    servable from a process that also has to answer other requests.
    """

    def get(self, request, pk):
        client = resolve_client(request)
        record = owned_uploads(request, client).filter(pk=pk).first()
        if not record:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        try:
            handle = record.stored_file.open("rb")
        except (OSError, ValueError) as exc:
            logger.warning("Could not open stored source for upload %s: %s", pk, exc)
            return Response(
                {"detail": "The stored source file could not be read."},
                status=status.HTTP_404_NOT_FOUND,
            )

        response = FileResponse(
            handle,
            as_attachment=True,
            filename=record.original_filename,
            content_type="application/octet-stream",
        )
        # Set explicitly so a browser shows real progress instead of an
        # indeterminate spinner on a file that takes a minute to arrive.
        if record.file_size_bytes:
            response["Content-Length"] = str(record.file_size_bytes)
        response["X-Content-Type-Options"] = "nosniff"
        return response


# ===========================================================================
# Part 6.4 — the segment dictionary comes from the database.
# ===========================================================================


class SegmentDictionaryView(APIView):
    """
    Every 834 segment and element the mapping UI may offer.

    The dropdowns were a literal in a React file, so the set of segments a user
    could map was whatever somebody had typed into the front end — nine of them
    — while SegmentElement sat seeded and migrated in the database with the
    rest. Adding a segment meant a front-end release. It is a management
    command now, or a row.
    """

    def get(self, request):
        include_inactive = (request.query_params.get("include_inactive") or "").lower() in (
            "1",
            "true",
            "yes",
        )
        queryset = SegmentElement.objects.all()
        if not include_inactive:
            queryset = queryset.filter(is_active=True)

        segments: dict = {}
        for item in queryset.order_by("segment_name", "element_code"):
            bucket = segments.setdefault(item.segment_name, [])
            # The dictionary is unique on (segment, element, loop), so the same
            # element can legitimately appear under two loops. The dropdown
            # wants one entry per element.
            if any(entry["code"] == item.element_code for entry in bucket):
                continue
            position = item.element_code[len(item.segment_name):]
            bucket.append(
                {
                    # Hyphenated for display, canonical for the wire. Both are
                    # sent so the browser never has to derive one from the
                    # other and get it subtly wrong.
                    "id": "{seg}-{pos}".format(seg=item.segment_name, pos=position),
                    "code": item.element_code,
                    "name": item.description,
                    "loop_id": item.loop_id,
                    "data_type": item.data_type,
                }
            )

        return Response(
            {
                "segments": segments,
                "segment_names": sorted(segments.keys()),
                "count": sum(len(v) for v in segments.values()),
            }
        )


# ===========================================================================
# Part 5 — the Information section, from the database.
# ===========================================================================


class DashboardSummaryView(APIView):
    """
    The counts behind the Information panel.

    Every number here was a literal in the markup. A dashboard whose figures do
    not move is worse than no dashboard: it is read as a measurement, and it
    reports the same reassuring totals whether the last upload succeeded, failed
    or never happened.

    Counted in the database rather than in Python. The alternative — fetch the
    rows and len() them — is fine at demo scale and is a full table read per
    tile once a sponsor has thirty thousand members.
    """

    def get(self, request):
        from django.db.models import Count, Max, Q, Sum

        client = resolve_client(request)
        uploads = owned_uploads(request, client)

        counts = uploads.aggregate(
            total=Count("id"),
            validated=Count(
                "id", filter=Q(processing_status__in=VALIDATED_STATUSES)
            ),
            converted=Count("id", filter=Q(processing_status=ProcessingStatus.CONVERTED)),
            failed=Count("id", filter=Q(processing_status=ProcessingStatus.FAILED)),
            quarantined=Count(
                "id", filter=Q(processing_status=ProcessingStatus.QUARANTINED)
            ),
            pending=Count(
                "id",
                filter=Q(
                    processing_status__in=(
                        ProcessingStatus.UPLOADED,
                        ProcessingStatus.PENDING,
                        ProcessingStatus.VALIDATING,
                        ProcessingStatus.PARSING,
                        ProcessingStatus.CONVERTING,
                    )
                ),
            ),
            latest_upload=Max("uploaded_at"),
            latest_file_date=Max("file_date"),
            member_loops=Sum("member_loop_count"),
        )

        subscribers = scope_to_client(
            Subscriber.objects.filter(owner=request.user), client
        ).count()
        dependants = scope_to_client(
            Dependant.objects.filter(owner=request.user), client
        ).count()

        conversions = scope_to_client(
            ConversionHistory.objects.filter(owner=request.user), client
        )
        conversion_counts = conversions.aggregate(
            runs=Count("id"),
            failed_runs=Count("id", filter=Q(status=ConversionHistory.Status.FAILED)),
            rows=Sum("rows_written"),
        )

        latest_upload = counts["latest_upload"]

        return Response(
            {
                "total_files": counts["total"] or 0,
                "validated_files": counts["validated"] or 0,
                "converted_files": counts["converted"] or 0,
                "failed_files": counts["failed"] or 0,
                "quarantined_files": counts["quarantined"] or 0,
                "in_progress_files": counts["pending"] or 0,
                "total_subscribers": subscribers,
                "total_dependants": dependants,
                "total_members": subscribers + dependants,
                "total_member_loops": counts["member_loops"] or 0,
                "conversion_runs": conversion_counts["runs"] or 0,
                "failed_conversions": conversion_counts["failed_runs"] or 0,
                "rows_generated": conversion_counts["rows"] or 0,
                "latest_upload_at": latest_upload,
                "latest_upload_display": display_date(latest_upload),
                "latest_file_date": counts["latest_file_date"],
                "latest_file_date_display": display_date(counts["latest_file_date"]),
            }
        )
