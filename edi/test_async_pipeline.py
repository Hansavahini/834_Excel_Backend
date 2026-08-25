"""
Regression tests for the defects fixed in the background-processing pass.

Each test here corresponds to something that was observed failing against a
real client file, not to a hypothetical. The comments say what the failure
looked like from the user's side, because that is the thing a future change is
at risk of reintroducing - the mechanism can be rewritten freely as long as the
symptom stays gone.

Jobs run inline under the test runner (settings.EDI_RUN_JOBS_INLINE), so a
response here already carries the finished result. That is a property of the
test environment, not of the API: in production these endpoints return 202 and
the browser polls /api/edi/jobs/.
"""

from __future__ import annotations

import io
from datetime import timedelta

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone

from conversion.models import ConversionHistory
from edi.models import JobKind, JobState, ProcessingJob
from edi.services.runner import reap_stale
from files.models import ProcessingStatus, UploadedFile
from mapping.models import MappingDetail, MappingTemplate

User = get_user_model()


def interchange(*, member_last="SMITH", city="DAYTON") -> str:
    """A minimal but structurally valid 834: one subscriber, one dependent."""
    segments = [
        "ISA*00*          *00*          *ZZ*ONESMARTER     *ZZ*ABCHEALTHPLAN  "
        "*240401*1200*^*00501*000000001*0*P*:",
        "GS*BE*ONESMARTER*ABCHEALTHPLAN*20240401*1200*1*X*005010X220A1",
        "ST*834*0001*005010X220A1",
        "BGN*00*00000001*20240401*1200****4",
        "REF*38*GRP1001",
        "DTP*007*D8*20240401",
        "N1*P5*ABC HEALTH PLAN SPONSOR*FI*351234567",
        "INS*Y*18*030*XN*A***FT",
        "REF*0F*100000000",
        "REF*1L*GRP1001",
        "NM1*IL*1*{last}*JOHN****34*100000000".format(last=member_last),
        "PER*IP**HP*9375550000*EM*john@example.com",
        "N3*100 MAIN ST",
        "N4*{city}*OH*45402".format(city=city),
        "DMG*D8*19600115*M",
        "HD*030**HLT*PPO-GOLD*IND",
        "DTP*348*D8*20240101",
        "INS*N*01*030*XN*A***FT",
        "REF*0F*100000000",
        "NM1*IL*1*{last}*LINDA****34*100000001".format(last=member_last),
        "DMG*D8*19620120*F",
        "HD*030**HLT*PPO-GOLD*IND",
        "DTP*348*D8*20240101",
    ]
    # SE01 counts ST through SE inclusive: everything after ISA and GS, plus
    # the SE segment itself.
    segments.append("SE*{n}*0001".format(n=len(segments) - 1))
    segments.append("GE*1*1")
    segments.append("IEA*1*000000001")
    return "~".join(segments) + "~"


RULES = [
    {"excel_column": "LAST NAME", "segment": "NM1", "element": "NM103",
     "qualifier_element": "NM101", "qualifier_value": "IL", "applies_to": "SUB"},
    {"excel_column": "FIRST NAME", "segment": "NM1", "element": "NM104",
     "qualifier_element": "NM101", "qualifier_value": "IL", "applies_to": "SUB"},
    {"excel_column": "CITY", "segment": "N4", "element": "N401", "applies_to": "SUB"},
    {"excel_column": "PHONE", "segment": "PER", "element": "PER04",
     "qualifier_element": "PER03", "qualifier_value": "HP", "applies_to": "SUB"},
]
COLUMNS = [rule["excel_column"] for rule in RULES]


class PipelineTestCase(TestCase):
    def setUp(self):
        self.user = User.objects.create_user("tester", password="tester12345")
        self.client.force_login(self.user)

    def upload(self, content=None, name="t.x12"):
        payload = (content or interchange()).encode("utf-8")
        response = self.client.post(
            reverse("upload"),
            {"file": SimpleUploadedFile(name, payload, content_type="text/plain")},
        )
        self.assertIn(response.status_code, (200, 201), response.content)
        return response.json()["uploaded_file_id"]

    def validate(self, file_id):
        return self.client.post(
            reverse("validate"),
            {"uploaded_file_id": file_id},
            content_type="application/json",
        )

    def convert(self, file_id, rules=None, columns=None, force=False):
        body = {
            "uploaded_file_id": file_id,
            "columns": columns or COLUMNS,
            "mappings": rules or RULES,
        }
        if force:
            body["force"] = True
        return self.client.post(
            reverse("convert"), body, content_type="application/json"
        )


class NoRequestBlocksOnTheWork(PipelineTestCase):
    """
    The reported symptom: "it gets stuck on validating, and once the page is
    refreshed, it shows the file is validated."

    The work was being done inside the request. On a real interchange that is
    minutes, so the browser waited on a connection something upstream eventually
    cut - while the server finished normally, which is why a refresh revealed
    the answer. Both endpoints must now hand back a job rather than a result.
    """

    def test_validate_creates_a_job_against_the_file(self):
        file_id = self.upload()
        payload = self.validate(file_id).json()

        self.assertIn("job_id", payload)
        job = ProcessingJob.objects.get(pk=payload["job_id"])
        self.assertEqual(job.kind, JobKind.VALIDATE)
        self.assertEqual(job.uploaded_file_id, file_id)

    def test_convert_creates_a_job_against_the_file(self):
        file_id = self.upload()
        self.validate(file_id)
        payload = self.convert(file_id).json()

        job = ProcessingJob.objects.get(pk=payload["job_id"])
        self.assertEqual(job.kind, JobKind.CONVERT)

    def test_a_second_press_does_not_queue_a_second_run(self):
        """
        Two clicks on Validate must not start two member syncs over one file.
        Doing so is not merely wasteful: the two passes interleave writes to the
        same member rows.
        """
        file_id = self.upload()
        self.validate(file_id)
        UploadedFile.objects.filter(pk=file_id).update(
            processing_status=ProcessingStatus.UPLOADED
        )
        ProcessingJob.objects.filter(uploaded_file_id=file_id).update(
            state=JobState.RUNNING, finished_at=None
        )

        again = self.validate(file_id)
        self.assertEqual(again.status_code, 200)
        self.assertEqual(
            ProcessingJob.objects.filter(
                uploaded_file_id=file_id, kind=JobKind.VALIDATE
            ).count(),
            1,
        )


class NoFileIsEverStranded(PipelineTestCase):
    """
    The reported symptom: a 500 from convert, and a row left reading CONVERTING
    for ever with a workbook name beside it.

    The conversion path caught (EDIParseError, ValueError, OSError) only.
    openpyxl's IllegalCharacterError inherits from none of them, so it escaped
    as a 500 after the GeneratedFile row had already been written - leaving a
    status with no exit and a filename that made it look as though something had
    succeeded.
    """

    def test_a_worker_that_dies_releases_the_file(self):
        file_id = self.upload()
        self.validate(file_id)
        self.convert(file_id)

        # Simulate the process being killed mid-conversion: the file sits at
        # CONVERTING and the job's heartbeat stops.
        UploadedFile.objects.filter(pk=file_id).update(
            processing_status=ProcessingStatus.CONVERTING
        )
        stale = timezone.now() - timedelta(hours=2)
        ProcessingJob.objects.filter(
            uploaded_file_id=file_id, kind=JobKind.CONVERT
        ).update(
            state=JobState.RUNNING, started_at=stale, heartbeat_at=stale, finished_at=None
        )

        self.assertEqual(reap_stale(), 1)

        record = UploadedFile.objects.get(pk=file_id)
        self.assertNotEqual(record.processing_status, ProcessingStatus.CONVERTING)
        job = ProcessingJob.objects.filter(
            uploaded_file_id=file_id, kind=JobKind.CONVERT
        ).first()
        self.assertEqual(job.state, JobState.INTERRUPTED)
        self.assertIn("run it again", job.error)

    def test_a_failing_conversion_leaves_the_file_convertible(self):
        """A failure must be recoverable by pressing the button again."""
        file_id = self.upload()
        self.validate(file_id)

        # A rule pointing at a segment the file does not contain is harmless;
        # to force a real failure, remove the stored source from under the run.
        record = UploadedFile.objects.get(pk=file_id)
        import os

        os.remove(record.stored_file.path)

        response = self.convert(file_id)
        self.assertIn(response.status_code, (200, 202))

        record.refresh_from_db()
        self.assertNotEqual(record.processing_status, ProcessingStatus.CONVERTING)
        self.assertTrue(record.conversion_error)


class ControlCharactersDoNotBreakTheWorkbook(PipelineTestCase):
    """
    Real 834s carry characters XML cannot represent: NUL padding from fixed
    width extracts, 0x1C/0x1D/0x1E used as X12 separators (which the standard
    permits), and anything at all after an EBCDIC conversion. openpyxl rejects
    them at save() time.
    """

    def test_a_file_with_control_characters_converts(self):
        raw = interchange(member_last="SM\x00IT\x1fH", city="DAY\x0bTON")
        file_id = self.upload(raw)
        self.validate(file_id)
        response = self.convert(file_id)
        self.assertIn(response.status_code, (200, 202), response.content)

        record = UploadedFile.objects.get(pk=file_id)
        self.assertEqual(record.processing_status, ProcessingStatus.CONVERTED)

    def test_the_workbook_opens_and_the_value_is_still_legible(self):
        from openpyxl import load_workbook

        raw = interchange(member_last="SM\x00IT\x1fH")
        file_id = self.upload(raw)
        self.validate(file_id)
        self.convert(file_id)

        generated = UploadedFile.objects.get(pk=file_id).generated_files.first()
        workbook = load_workbook(generated.stored_file.path, read_only=True)
        try:
            rows = list(workbook.worksheets[0].iter_rows(values_only=True))
        finally:
            workbook.close()

        names = [row[COLUMNS.index("LAST NAME")] for row in rows[1:]]
        # Substituted, not deleted: running the fragments together would make
        # a corrupted value look like a real surname.
        self.assertTrue(any("SM IT H" == value for value in names), names)


class MappingVersionsAreNotInflated(PipelineTestCase):
    """
    A completed conversion locks the template version it used, so the next edit
    clones to version n+1. Correct - it is what keeps history true. But Convert
    saves the mapping before every run whether or not anything was edited, so
    three identical clicks produced four versions, each a full copy of every
    rule, growing the tables without a single rule ever changing.
    """

    def test_repeated_conversions_with_an_unchanged_mapping_mint_no_versions(self):
        file_id = self.upload()
        self.validate(file_id)
        self.convert(file_id)

        templates = MappingTemplate.objects.count()
        details = MappingDetail.objects.count()

        for _ in range(3):
            self.convert(file_id)

        self.assertEqual(MappingTemplate.objects.count(), templates)
        self.assertEqual(MappingDetail.objects.count(), details)

    def test_an_actual_mapping_change_does_mint_a_version(self):
        file_id = self.upload()
        self.validate(file_id)
        self.convert(file_id)
        before = MappingTemplate.objects.count()

        changed = [dict(rule) for rule in RULES]
        changed[2] = {**changed[2], "element": "N402"}  # CITY now reads the state
        self.convert(file_id, rules=changed)

        self.assertEqual(MappingTemplate.objects.count(), before + 1)


class AMappingChangeReachesConvertedFiles(PipelineTestCase):
    """
    "When I am changing the mapping that mapping should get applied to the
    converted files." It must, and it must not silently rebuild files whose
    mapping did not change - one click on a client with thirty files was thirty
    conversions producing thirty identical workbooks.
    """

    def test_the_new_mapping_produces_a_new_workbook(self):
        from openpyxl import load_workbook

        file_id = self.upload()
        self.validate(file_id)
        self.convert(file_id)

        changed = [dict(rule) for rule in RULES]
        changed[2] = {**changed[2], "element": "N402"}
        self.convert(file_id, rules=changed)

        generated = (
            UploadedFile.objects.get(pk=file_id)
            .generated_files.order_by("-generated_at")
            .first()
        )
        workbook = load_workbook(generated.stored_file.path, read_only=True)
        try:
            rows = list(workbook.worksheets[0].iter_rows(values_only=True))
        finally:
            workbook.close()

        cities = [row[COLUMNS.index("CITY")] for row in rows[1:]]
        self.assertTrue(all(value == "OH" for value in cities), cities)

    def test_an_unchanged_mapping_reuses_the_existing_workbook(self):
        file_id = self.upload()
        self.validate(file_id)
        self.convert(file_id)
        first = UploadedFile.objects.get(pk=file_id).generated_files.count()

        payload = self.convert(file_id).json()
        self.assertTrue(payload.get("result", payload).get("skipped"))
        self.assertEqual(
            UploadedFile.objects.get(pk=file_id).generated_files.count(), first
        )

    def test_force_rebuilds_even_when_the_mapping_is_unchanged(self):
        file_id = self.upload()
        self.validate(file_id)
        self.convert(file_id)
        first = UploadedFile.objects.get(pk=file_id).generated_files.count()

        self.convert(file_id, force=True)
        self.assertEqual(
            UploadedFile.objects.get(pk=file_id).generated_files.count(), first + 1
        )


class UnmappingAColumnActuallyUnmapsIt(PipelineTestCase):
    """
    Setting a row's segment back to "Select..." looked like it removed the
    mapping. It did not: the save path only ever wrote rows, so the
    MappingDetail survived and the workbook kept filling a column the screen
    showed as unmapped.
    """

    def test_a_removed_rule_no_longer_fills_the_column(self):
        self.client.post(
            reverse("mappings"),
            {"columns": COLUMNS, "mappings": RULES},
            content_type="application/json",
        )
        remaining = [rule for rule in RULES if rule["excel_column"] != "PHONE"]
        self.client.post(
            reverse("mappings"),
            {"columns": COLUMNS, "mappings": remaining},
            content_type="application/json",
        )

        saved = self.client.get(reverse("mappings")).json()
        phone = next(c for c in saved["columns"] if c["excel_column"] == "PHONE")
        self.assertEqual(phone["segment"], "")
        self.assertEqual(phone["element"], "")

    def test_the_column_itself_survives_being_unmapped(self):
        """
        Un-mapping must narrow the rule, never delete the column.

        The first cut of this fix deleted the MappingDetail and nothing else,
        so the column vanished from the grid the screen rebuilds after every
        save - leaving no way to map it again short of Reset, which discards
        every other customisation on the screen.
        """
        self.client.post(
            reverse("mappings"),
            {"columns": COLUMNS, "mappings": RULES},
            content_type="application/json",
        )
        remaining = [rule for rule in RULES if rule["excel_column"] != "PHONE"]
        self.client.post(
            reverse("mappings"),
            {"columns": COLUMNS, "mappings": remaining},
            content_type="application/json",
        )

        saved = self.client.get(reverse("mappings")).json()
        columns = [column["excel_column"] for column in saved["columns"]]
        self.assertEqual(columns, COLUMNS)

    def test_an_unmapped_column_still_appears_in_the_workbook(self):
        """LOCAL and CLASS are filled in downstream by hand; they must be there."""
        from openpyxl import load_workbook

        file_id = self.upload()
        self.validate(file_id)
        self.convert(file_id, columns=COLUMNS + ["LOCAL"])

        generated = UploadedFile.objects.get(pk=file_id).generated_files.first()
        workbook = load_workbook(generated.stored_file.path, read_only=True)
        try:
            header = next(workbook.worksheets[0].iter_rows(values_only=True))
        finally:
            workbook.close()

        self.assertIn("LOCAL", header)


class TheAuditTrailStaysTrue(PipelineTestCase):
    """
    A conversion must be able to say which rules produced it, and the column
    layout is part of that answer: adding an unmapped column changes the
    workbook without changing a single rule.
    """

    def test_history_records_the_rules_and_the_layout(self):
        file_id = self.upload()
        self.validate(file_id)
        self.convert(file_id, columns=COLUMNS + ["LOCAL"])

        history = ConversionHistory.objects.filter(uploaded_file_id=file_id).first()
        self.assertEqual(history.status, ConversionHistory.Status.SUCCESS)
        self.assertIn("LOCAL", history.result_headers)
        self.assertEqual(len(history.mapping_snapshot), len(RULES))


class PreviewsArePaged(PipelineTestCase):
    """
    Both viewers used to pull as much as the endpoint would give and render all
    of it. The endpoints must page, so a large file opens instantly instead of
    locking the tab.
    """

    def test_the_workbook_preview_takes_an_offset(self):
        file_id = self.upload()
        self.validate(file_id)
        self.convert(file_id)
        generated = UploadedFile.objects.get(pk=file_id).generated_files.first()

        whole = self.client.get(
            reverse("download-preview", args=[generated.id]), {"limit": 50, "offset": 0}
        ).json()
        page = self.client.get(
            reverse("download-preview", args=[generated.id]), {"limit": 50, "offset": 1}
        ).json()

        self.assertGreater(len(whole["rows"]), 1)
        self.assertEqual(page["offset"], 1)
        # Offsetting by one drops exactly the first row and keeps the rest, so
        # the two pages line up. Comparing row contents would not prove this:
        # a subscriber and their dependent legitimately carry identical values
        # in the subscriber-scoped columns.
        self.assertEqual(page["rows"], whole["rows"][1:])
        self.assertFalse(page["has_more"])

    def test_the_source_preview_takes_an_offset(self):
        file_id = self.upload()
        head = self.client.get(
            reverse("edi-file-preview", args=[file_id]), {"limit": 40, "offset": 0}
        ).json()
        tail = self.client.get(
            reverse("edi-file-preview", args=[file_id]), {"limit": 40, "offset": 40}
        ).json()

        self.assertTrue(head["truncated"])
        self.assertNotEqual(head["content"], tail["content"])

    def test_the_source_download_streams_with_a_length(self):
        """
        The download must advertise its size so the browser shows real progress
        rather than an indeterminate spinner - and so the front end never needs
        to buffer the file to know how big it is.
        """
        file_id = self.upload()
        response = self.client.get(reverse("edi-file-download", args=[file_id]))

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response["Content-Length"])
        self.assertIn("attachment", response["Content-Disposition"])


class DownloadsAreCountedOnce(PipelineTestCase):
    """
    downloaded_count is the record of who took PHI out of the system and how
    often. The browser download path issues a HEAD first so a 403 or 404
    surfaces as a message rather than an error page inside a hidden iframe -
    and Django maps HEAD onto the GET handler unless a view defines its own, so
    that preflight was recording a second download every time.
    """

    def test_a_head_preflight_is_not_a_download(self):
        file_id = self.upload()
        self.validate(file_id)
        self.convert(file_id)
        generated = UploadedFile.objects.get(pk=file_id).generated_files.first()

        self.client.head(reverse("download", args=[generated.id]))
        generated.refresh_from_db()
        self.assertEqual(generated.downloaded_count, 0)

        self.client.get(reverse("download", args=[generated.id]))
        generated.refresh_from_db()
        self.assertEqual(generated.downloaded_count, 1)

    def test_head_still_reports_the_size_and_filename(self):
        file_id = self.upload()
        self.validate(file_id)
        self.convert(file_id)
        generated = UploadedFile.objects.get(pk=file_id).generated_files.first()

        response = self.client.head(reverse("download", args=[generated.id]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Length"], str(generated.file_size_bytes)
        )
        self.assertIn(generated.generated_filename, response["Content-Disposition"])

    def test_head_on_someone_elses_file_is_not_found(self):
        """The preflight must not become a way to probe for other people's files."""
        file_id = self.upload()
        self.validate(file_id)
        self.convert(file_id)
        generated = UploadedFile.objects.get(pk=file_id).generated_files.first()

        intruder = User.objects.create_user("intruder", password="intruder12345")
        self.client.force_login(intruder)
        response = self.client.head(reverse("download", args=[generated.id]))
        self.assertEqual(response.status_code, 404)

    def test_the_source_download_head_does_not_open_the_file(self):
        file_id = self.upload()
        record = UploadedFile.objects.get(pk=file_id)
        import os

        os.remove(record.stored_file.path)

        # HEAD answers from the row, so a missing file on disk is still a
        # well-formed header response rather than a 404 the user cannot action.
        response = self.client.head(reverse("edi-file-download", args=[file_id]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Length"], str(record.file_size_bytes))
