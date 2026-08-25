"""
The acceptance tests from the brief, as tests rather than as a one-off script.

Each class names the part of the brief it covers. The point of keeping them
here rather than in a scratch file is that every one of these describes a defect
that was shipped once; a test is the only thing that stops it being shipped
twice.
"""

from __future__ import annotations

import json
import tempfile

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings

from conversion.models import ConversionHistory
from files.models import GeneratedFile, ProcessingStatus, UploadedFile
from members.models import Dependant, EnrollmentRecord, Subscriber

# A minimal but structurally valid 834: one subscriber, one child, both with
# their SSN in NM109 under NM108=34 and a different value in REF*0F, which is
# the arrangement that made the old extraction look correct on subscribers and
# be wrong on dependants.
SEGMENTS = [
    "ISA*00*          *00*          *ZZ*ONESMARTER     *ZZ*ABCHEALTHPLAN  *240401*1200*^*00501*000000001*0*P*:",
    "GS*BE*ONESMARTER*ABCHEALTHPLAN*{date}*1200*1*X*005010X220A1",
    "ST*834*0001*005010X220A1",
    "BGN*00*00000001*{date}*1200****4",
    "N1*P5*ABC HEALTH PLAN SPONSOR*FI*351234567",
    "INS*Y*18*030*XN*A***FT",
    "REF*0F*SUBNUM777",
    "REF*1L*GRP1001",
    "NM1*IL*1*ZEROTEST*ALICE****34*001234567",
    "DMG*D8*19700304*F",
    "HD*030**HLT*PPO-GOLD*IND",
    "DTP*348*D8*{eff}",
    "INS*N*19*030*XN*A***FT",
    "REF*0F*SUBNUM777",
    "NM1*IL*1*ZEROTEST*BOBBY****34*000998877",
    "DMG*D8*20120909*M",
    "HD*030**HLT*PPO-GOLD*IND",
    "DTP*348*D8*{eff}",
    "SE*17*0001",
    "GE*1*1",
    "IEA*1*000000001",
]

RULES = [
    {
        "excel_column": "LAST NAME",
        "segment": "NM1",
        "element": "NM103",
        "qualifier_element": "NM101",
        "qualifier_value": "IL",
        "column_order": 1,
    },
    {
        "excel_column": "SSN",
        "segment": "NM1",
        "element": "NM109",
        "qualifier_element": "NM108",
        "qualifier_value": "34",
        "transform": "SSN",
        "column_order": 2,
    },
    {
        "excel_column": "DOB",
        "segment": "DMG",
        "element": "DMG02",
        "transform": "DATE_MDY",
        "column_order": 3,
    },
    {
        "excel_column": "EFF DATE",
        "segment": "DTP",
        "element": "DTP03",
        "qualifier_element": "DTP01",
        "qualifier_value": "348",
        "transform": "DATE_MDY",
        "column_order": 4,
    },
]


def build_834(date="20250101", eff="20250101"):
    body = "~".join(seg.format(date=date, eff=eff) for seg in SEGMENTS) + "~"
    return body.encode()


class BriefTestCase(TestCase):
    """Shared plumbing. MEDIA_ROOT is per-test so nothing leaks between them."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="brief", password="brief-password", is_staff=True
        )

    def setUp(self):
        self.client.force_login(self.user)
        self._media = tempfile.TemporaryDirectory()
        self.addCleanup(self._media.cleanup)
        override = override_settings(MEDIA_ROOT=self._media.name)
        override.enable()
        self.addCleanup(override.disable)

    def upload(self, name="brief.x12", date="20250101", eff="20250101"):
        """
        Upload and validate, as the portal now does in two steps.

        Upload stores bytes and returns immediately; validation, envelope
        facts and the member sync run when Validate is pressed. Tests that
        exercise post-validation behaviour therefore go through both, and the
        returned response is the upload response the callers already unpack.
        """
        created = self.client.post(
            "/api/edi/upload/",
            {"file": SimpleUploadedFile(name, build_834(date, eff))},
        )
        body = created.json()
        if created.status_code == 201 and body.get("uploaded_file_id"):
            self.client.post(
                "/api/edi/validate/",
                json.dumps({"uploaded_file_id": body["uploaded_file_id"]}),
                content_type="application/json",
            )
        return created

    def convert(self, uploaded_file_id, rules=None):
        rules = RULES if rules is None else rules
        return self.client.post(
            "/api/edi/convert/",
            json.dumps(
                {
                    "uploaded_file_id": uploaded_file_id,
                    "columns": [rule["excel_column"] for rule in rules],
                    "mappings": rules,
                }
            ),
            content_type="application/json",
        )

    def listing_entry(self, uploaded_file_id):
        listing = self.client.get("/api/edi/uploads/").json()
        return next(item for item in listing if item["id"] == uploaded_file_id)


class Part1StatusSurvivesRefresh(BriefTestCase):
    """Test 1 and Test 2: a browser refresh must not undo completed work."""

    def test_validation_status_is_stored(self):
        created = self.upload().json()
        self.client.post(
            "/api/edi/validate/",
            json.dumps({"uploaded_file_id": created["uploaded_file_id"]}),
            content_type="application/json",
        )

        entry = self.listing_entry(created["uploaded_file_id"])
        self.assertEqual(entry["status"], ProcessingStatus.VALIDATED)
        self.assertIsNotNone(entry["validated_at"])
        # MM-DD-YYYY, formatted once on the server.
        self.assertRegex(entry["validated_at_display"], r"^\d{2}-\d{2}-\d{4}$")

    def test_conversion_status_and_download_are_stored(self):
        created = self.upload().json()
        response = self.convert(created["uploaded_file_id"])
        self.assertEqual(response.status_code, 200)

        entry = self.listing_entry(created["uploaded_file_id"])
        self.assertEqual(entry["status"], ProcessingStatus.CONVERTED)
        self.assertTrue(entry["is_converted"])
        self.assertIsNotNone(entry["converted_at"])
        self.assertIsNotNone(entry["generated_file_id"])

        # The download link restored from the database actually serves the file.
        self.assertEqual(self.client.get(entry["download_url"]).status_code, 200)

    def test_a_failed_conversion_does_not_strand_the_file(self):
        """A status nothing can leave is how a file becomes permanently unusable."""
        created = self.upload().json()
        response = self.convert(
            created["uploaded_file_id"],
            rules=[{"excel_column": "X", "segment": "NM1", "element": "NM103", "column_order": 1}],
        )
        self.assertEqual(response.status_code, 200)
        record = UploadedFile.objects.get(pk=created["uploaded_file_id"])
        self.assertIn(
            record.processing_status,
            (ProcessingStatus.CONVERTED, ProcessingStatus.VALIDATED),
        )


class Part2And3MasterTables(BriefTestCase):
    """Tests 3, 4 and 6: one row per person per role, history preserved."""

    def test_same_file_twice_creates_no_duplicate_masters(self):
        self.upload(name="one.x12")
        subscribers = Subscriber.objects.count()
        dependants = Dependant.objects.count()

        again = self.upload(name="one.x12").json()

        self.assertTrue(again["duplicate"])
        self.assertEqual(Subscriber.objects.count(), subscribers)
        self.assertEqual(Dependant.objects.count(), dependants)

    def test_same_ssn_in_two_files_is_one_master_and_two_enrollments(self):
        self.upload(name="jan.x12", date="20250101", eff="20250101")
        self.upload(name="jan26.x12", date="20260101", eff="20260101")

        subscriber = Subscriber.objects.get(ssn="001234567")
        self.assertEqual(Subscriber.objects.filter(ssn="001234567").count(), 1)

        # History is not overwritten: one enrollment per file, both readable.
        enrollments = list(subscriber.enrollments.order_by("effective_date"))
        self.assertEqual(len(enrollments), 2)
        self.assertEqual(
            [str(row.effective_date) for row in enrollments],
            ["2025-01-01", "2026-01-01"],
        )

    def test_the_two_tables_hold_only_their_own_role(self):
        self.upload()

        self.assertEqual(
            Subscriber.objects.filter(source_member__member_type="DEP").count(), 0
        )
        self.assertEqual(
            Dependant.objects.filter(source_member__member_type="SUB").count(), 0
        )

        dependant = Dependant.objects.get(ssn="000998877")
        self.assertIsNotNone(dependant.subscriber_id)
        self.assertEqual(dependant.subscriber.ssn, "001234567")


class Part14SsnExtraction(BriefTestCase):
    """
    The defect worth naming twice: REF*0F is the subscriber number.

    Both people in the fixture carry REF*0F*SUBNUM777. Reading that as an SSN
    gave the child the subscriber's number, and de-duplicating on SSN would then
    have merged them into one person.
    """

    def test_ref_0f_is_never_read_as_an_ssn(self):
        self.upload()

        self.assertFalse(Subscriber.objects.filter(ssn="SUBNUM777").exists())
        self.assertEqual(Subscriber.objects.get(ssn="001234567").last_name, "ZEROTEST")
        self.assertEqual(Dependant.objects.get(ssn="000998877").first_name, "BOBBY")

    def test_the_subscriber_number_is_still_recorded(self):
        """Not read as an SSN, but not discarded either."""
        self.upload()
        subscriber = Subscriber.objects.get(ssn="001234567")
        self.assertEqual(subscriber.source_member.subscriber_number, "SUBNUM777")


class Part4And14ExcelOutput(BriefTestCase):
    """Test 5: date cells and text SSNs, in the workbook rather than as strings."""

    def workbook(self):
        created = self.upload().json()
        self.convert(created["uploaded_file_id"])
        generated = GeneratedFile.objects.order_by("-generated_at").first()
        from openpyxl import load_workbook

        return load_workbook(generated.stored_file.path).worksheets[0]

    def test_dates_are_real_date_cells_formatted_mm_dd_yyyy(self):
        sheet = self.workbook()
        headers = [cell.value for cell in sheet[1]]
        column = headers.index("DOB") + 1

        cell = sheet.cell(row=2, column=column)
        self.assertEqual(cell.number_format, "MM-DD-YYYY")
        # A real date, not a string that looks like one.
        self.assertTrue(hasattr(cell.value, "year"))

    def test_ssn_keeps_nine_digits_and_its_leading_zeros(self):
        sheet = self.workbook()
        headers = [cell.value for cell in sheet[1]]
        column = headers.index("SSN") + 1

        values = [
            (sheet.cell(row=row, column=column).value,
             sheet.cell(row=row, column=column).number_format)
            for row in range(2, sheet.max_row + 1)
        ]

        for value, number_format in values:
            self.assertIsInstance(value, str)
            self.assertEqual(len(value), 9)
            # Text format is what actually stops Excel showing 1234567.
            self.assertEqual(number_format, "@")

        self.assertTrue(any(value.startswith("00") for value, _ in values))


class Part13MemberSection(BriefTestCase):
    """Member search: SSN, name, validation errors and per-file history."""

    def test_leading_zero_ssn_search_finds_the_right_person(self):
        self.upload()
        response = self.client.get("/api/members/search/", {"q": "001234567"})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        # The screen shows the whole family: the matched subscriber first,
        # then their dependants, selected by radio button. So the hit widens
        # to everyone sharing that subscriber rather than the one row.
        self.assertEqual(payload[0]["last_name"], "ZEROTEST")
        self.assertEqual(payload[0]["record_type"], "SUBSCRIBER")
        matches = [row for row in payload if row["ssn_last4"] == "4567"]
        self.assertEqual(len(matches), 1)
        head = payload[0]["id"]
        for row in payload[1:]:
            self.assertEqual(row["subscriber_id"], head)

    def test_a_punctuated_ssn_is_the_same_person(self):
        self.upload()
        response = self.client.get("/api/members/search/", {"q": "001-23-4567"})
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        # Same family as the unpunctuated spelling, resolved through the
        # fingerprint: exactly one row carries the searched number.
        self.assertEqual(
            len([row for row in payload if row["ssn_last4"] == "4567"]), 1
        )
        self.assertEqual(payload[0]["record_type"], "SUBSCRIBER")

    def test_a_short_ssn_is_a_validation_error_not_an_empty_result(self):
        """
        "No member found" for 12345 is a lie by omission. There is no member
        because 12345 cannot be an SSN, and those are different problems.
        """
        self.upload()
        response = self.client.get("/api/members/search/", {"q": "12345"})

        self.assertEqual(response.status_code, 400)
        self.assertIn("nine", response.json()["detail"].lower())

    def test_name_search_works(self):
        self.upload()
        response = self.client.get("/api/members/search/", {"q": "zerotest"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.json()), 2)

    def test_history_shows_every_file_with_display_dates(self):
        self.upload(name="jan.x12", date="20250101", eff="20250101")
        self.upload(name="jan26.x12", date="20260101", eff="20260101")

        payload = self.client.get("/api/members/search/", {"q": "001234567"}).json()
        history = payload[0]["enrollment_history"]

        self.assertEqual(len(history), 2)
        self.assertEqual(
            sorted(row["effective_date_display"] for row in history),
            ["01-01-2025", "01-01-2026"],
        )
        self.assertEqual(
            sorted(row["file_name"] for row in history), ["jan.x12", "jan26.x12"]
        )

    def test_the_api_never_sends_a_plaintext_ssn(self):
        self.upload()
        payload = self.client.get("/api/members/search/", {"q": "001234567"}).json()

        self.assertNotIn("ssn", payload[0])
        self.assertEqual(payload[0]["masked_ssn"], "XXX-XX-4567")

    def test_the_separated_tables_are_readable_on_their_own(self):
        self.upload()

        subscribers = self.client.get("/api/members/subscribers/").json()
        dependants = self.client.get("/api/members/dependants/").json()

        self.assertEqual(subscribers["count"], 1)
        self.assertEqual(dependants["count"], 1)
        self.assertEqual(subscribers["results"][0]["record_type"], "SUBSCRIBER")
        self.assertEqual(dependants["results"][0]["record_type"], "DEPENDANT")
        # Masked in the list view too; a roster page renders hundreds of rows.
        self.assertEqual(subscribers["results"][0]["masked_ssn"], "XXX-XX-4567")


class Part5DashboardIsDynamic(BriefTestCase):
    def test_counts_come_from_the_database(self):
        created = self.upload().json()
        self.convert(created["uploaded_file_id"])

        summary = self.client.get("/api/edi/dashboard/").json()

        self.assertEqual(summary["total_files"], 1)
        self.assertEqual(summary["converted_files"], 1)
        self.assertEqual(summary["total_subscribers"], 1)
        self.assertEqual(summary["total_dependants"], 1)
        self.assertEqual(summary["total_members"], 2)
        self.assertRegex(summary["latest_upload_display"], r"^\d{2}-\d{2}-\d{4}$")

    def test_an_empty_deployment_reports_zero_rather_than_a_placeholder(self):
        summary = self.client.get("/api/edi/dashboard/").json()
        self.assertEqual(summary["total_files"], 0)
        self.assertEqual(summary["total_members"], 0)
        self.assertEqual(summary["latest_upload_display"], "")


class Part6MappingSystem(BriefTestCase):
    def test_saved_mappings_are_returned_for_reload(self):
        """Issue 6.1: they were saved and then ignored on every page load."""
        self.client.post(
            "/api/edi/mappings/", json.dumps(RULES), content_type="application/json"
        )

        payload = self.client.get("/api/edi/mappings/").json()
        columns = {column["excel_column"] for column in payload["columns"]}

        self.assertIn("SSN", columns)
        self.assertIn("DOB", columns)
        self.assertIsNotNone(payload["template_id"])

    def test_conversion_records_the_exact_mapping_it_used(self):
        created = self.upload().json()
        self.convert(created["uploaded_file_id"])

        history = ConversionHistory.objects.latest("created_at")

        self.assertEqual(history.mapping_source, "INLINE")
        self.assertEqual(len(history.mapping_snapshot), len(RULES))
        # Inline rules are persisted, so the run has a template and a version
        # rather than recording nothing at all.
        self.assertIsNotNone(history.mapping_template_id)
        self.assertIsNotNone(history.mapping_version)

        snapshot = {row["excel_column"]: row for row in history.mapping_snapshot}
        self.assertEqual(snapshot["SSN"]["qualifier_element"], "NM108")
        self.assertEqual(snapshot["SSN"]["qualifier_value"], "34")

    def test_the_segment_dictionary_is_served_from_the_database(self):
        """Issue 6.4: it was a literal in a React component."""
        from django.core.management import call_command

        call_command("seed_segment_elements", verbosity=0)
        payload = self.client.get("/api/edi/segments/").json()

        self.assertIn("NM1", payload["segments"])
        self.assertIn("HD", payload["segments"])
        # Wider than the nine segments the front end used to hard-code.
        self.assertGreater(len(payload["segment_names"]), 9)

    def test_occurrence_is_accepted_and_stored(self):
        """Issue 6.3: the resolver honoured it; nothing could set it."""
        rules = [dict(RULES[0], occurrence=2)]
        response = self.client.post(
            "/api/edi/mappings/", json.dumps(rules), content_type="application/json"
        )

        self.assertEqual(response.status_code, 201)
        payload = self.client.get("/api/edi/mappings/").json()
        column = next(c for c in payload["columns"] if c["excel_column"] == "LAST NAME")
        self.assertEqual(column["occurrence"], 2)


class Part8And9FileHandling(BriefTestCase):
    def test_download_returns_the_complete_file(self):
        created = self.upload().json()
        record = UploadedFile.objects.get(pk=created["uploaded_file_id"])

        response = self.client.get("/api/edi/files/{pk}/download/".format(pk=record.pk))
        body = b"".join(response.streaming_content)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(body), record.file_size_bytes)
        self.assertEqual(body, build_834())

    def test_preview_is_a_separate_endpoint_and_reports_truncation(self):
        created = self.upload().json()
        record = UploadedFile.objects.get(pk=created["uploaded_file_id"])

        response = self.client.get(
            "/api/edi/files/{pk}/preview/".format(pk=record.pk), {"limit": 40}
        )
        payload = response.json()

        self.assertEqual(len(payload["content"]), 40)
        self.assertTrue(payload["truncated"])
        self.assertIn("/download/", payload["download_url"])

    def test_another_users_file_is_not_reachable(self):
        created = self.upload().json()
        other = get_user_model().objects.create_user(
            username="intruder", password="intruder-password"
        )
        self.client.force_login(other)

        response = self.client.get(
            "/api/edi/files/{pk}/download/".format(pk=created["uploaded_file_id"])
        )
        self.assertEqual(response.status_code, 404)

    @override_settings(MAX_834_UPLOAD_BYTES=64)
    def test_the_upload_size_limit_is_enforced(self):
        """Issue 9: MAX_834_UPLOAD_BYTES was a setting nothing read."""
        response = self.upload()

        self.assertEqual(response.status_code, 413)
        payload = response.json()
        self.assertEqual(payload["max_upload_bytes"], 64)
        self.assertEqual(UploadedFile.objects.count(), 0)
