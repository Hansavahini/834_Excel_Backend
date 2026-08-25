"""
Workbook writer.

Three problems with the original. It saved to a bare relative filename, so the
workbook landed in whatever directory the process happened to start in, which
under gunicorn is not a directory anyone can find. The filename was the constant
"converted_834.xlsx", so two users converting at the same time overwrote each
other and the second one silently got the first one's PHI. And it built the
whole sheet in memory, which on a 27,000 member file is a large object graph
held for the duration of a request.

This version writes into MEDIA_ROOT/generated under a unique name, uses
openpyxl's write-only mode so rows are streamed to the zip as they are produced,
and returns enough metadata to populate a GeneratedFile row.
"""

from __future__ import annotations

import logging
import os
import re
import uuid
from dataclasses import dataclass
from datetime import date, datetime
from typing import Iterable, List, Optional

from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .transforms import KIND_DATE, KIND_TEXT, to_date

logger = logging.getLogger("edi.excel")

# The one date format this portal writes. Overridable through the environment
# so a client who insists on something else is a settings change, not a patch.
EXCEL_DATE_FORMAT = getattr(settings, "EXCEL_DATE_FORMAT", "MM-DD-YYYY")

# Excel's text format. Without it, "001234567" is stored as text but still shown
# as 1234567 the moment the file is opened, which is the exact defect Part 14
# describes: nothing in the pipeline reports an error and the client receives a
# workbook with eight digit Social Security Numbers in it.
TEXT_FORMAT = "@"

MAX_SHEET_ROWS = 1_048_576  # Excel's hard limit, worth failing loudly on

# Excel's own hard limit on a cell. Longer than this and openpyxl writes a file
# Excel refuses to open, which is worse than a truncated cell with a marker.
MAX_CELL_CHARS = 32_767

# Characters XML 1.0 cannot represent, which openpyxl rejects with
# IllegalCharacterError. Real 834 files contain them: some sponsors pad fixed
# width extracts with NUL, some use 0x1C/0x1D/0x1E as the segment, element and
# component separators (the X12 standard permits it), and a file that has been
# through a mainframe EBCDIC conversion can carry almost anything in a name
# field. openpyxl raises at save() time, long after the last useful stack frame,
# and IllegalCharacterError does not inherit from ValueError - so the old
# converter's `except (EDIParseError, ValueError, OSError)` let it straight
# through as a 500 with the file stranded at CONVERTING.
_ILLEGAL_XML = re.compile(
    r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x84\x86-\x9f\ud800-\udfff\ufdd0-\ufddf\ufffe\uffff]"
)


def sanitize_cell(value):
    """
    Make one value safe to write, without pretending it was clean.

    Control characters are replaced with a space rather than deleted, so a name
    field padded with NUL reads as two words instead of silently running them
    together and looking like real data. Over-long values are cut at Excel's
    limit with a visible marker, for the same reason.
    """
    if value is None or isinstance(value, (int, float, bool, date, datetime)):
        return value

    text = value if isinstance(value, str) else str(value)
    if _ILLEGAL_XML.search(text):
        text = _ILLEGAL_XML.sub(" ", text)
    if len(text) > MAX_CELL_CHARS:
        text = text[: MAX_CELL_CHARS - 15] + "...[truncated]"
    return text


@dataclass
class GeneratedWorkbook:
    absolute_path: str
    relative_path: str
    filename: str
    row_count: int
    size_bytes: int


def _output_dir(owner_id=None) -> str:
    parts = [str(settings.MEDIA_ROOT), getattr(settings, "GENERATED_SUBDIR", "generated")]
    if owner_id:
        parts.append(str(owner_id))
    parts.append(datetime.now().strftime("%Y/%m/%d"))
    directory = os.path.join(*parts)
    os.makedirs(directory, exist_ok=True)
    return directory


def build_filename(source_name: str = "834") -> str:
    """Unique per run, and still recognisable to whoever downloads it."""
    stem = os.path.splitext(os.path.basename(source_name or "834"))[0][:40]
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return "{stem}_{stamp}_{token}.xlsx".format(stem=stem, stamp=stamp, token=uuid.uuid4().hex[:8])


def _typed_cell(sheet, value, kind: str):
    """
    One cell, written as the thing it is rather than as a string that resembles it.

    DATE cells carry a real date and a number format, so the workbook sorts and
    filters by date and a recipient's Excel does not re-interpret the string
    under their own locale. TEXT cells carry the text format, which is what
    actually preserves a leading zero. Everything else is written as it arrives.
    """
    from openpyxl.cell import WriteOnlyCell

    if kind == KIND_DATE:
        parsed = to_date(value)
        if parsed is not None:
            cell = WriteOnlyCell(sheet, value=parsed)
            cell.number_format = EXCEL_DATE_FORMAT
            return cell
        # Unparseable. Write what the file said rather than a blank, so the
        # problem is visible to whoever has to fix the mapping.
        return WriteOnlyCell(sheet, value=sanitize_cell("" if value is None else str(value)))

    if kind == KIND_TEXT:
        cell = WriteOnlyCell(sheet, value=sanitize_cell("" if value is None else str(value)))
        cell.number_format = TEXT_FORMAT
        cell.alignment = Alignment(horizontal="left")
        return cell

    return sanitize_cell(value)


def generate_excel(
    headers: List[str],
    rows: Iterable[dict],
    output_path: Optional[str] = None,
    *,
    owner_id=None,
    source_name: str = "834",
    sheet_title: str = "834 Conversion",
    column_kinds: Optional[dict] = None,
) -> GeneratedWorkbook:
    """
    Write the workbook and return where it went.

    output_path is still honoured when a caller passes one, so existing tests
    keep working, but it is now resolved against MEDIA_ROOT when relative rather
    than against the process working directory.
    """
    if not headers:
        raise ValueError("Cannot generate a workbook with no columns; check the mapping template.")

    if output_path:
        if not os.path.isabs(output_path):
            output_path = os.path.join(_output_dir(owner_id), os.path.basename(output_path))
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        absolute_path = output_path
    else:
        absolute_path = os.path.join(_output_dir(owner_id), build_filename(source_name))

    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet(title=sheet_title[:31])

    bold = Font(bold=True)
    header_cells = []
    for text in headers:
        from openpyxl.cell import WriteOnlyCell

        cell = WriteOnlyCell(sheet, value=sanitize_cell(text))
        cell.font = bold
        header_cells.append(cell)
    sheet.append(header_cells)

    # Openpyxl's write-only sheets cannot set column widths after the fact, so
    # approximate from the header text; it is the difference between a readable
    # workbook and forty columns of '#####'.
    sheet.column_dimensions  # touch so the attribute exists
    for index, text in enumerate(headers, start=1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = min(max(len(str(text)) + 4, 12), 40)

    # header -> DATE | TEXT | GENERAL. Derived from the mapping transforms by
    # the caller, because only the mapping knows that "DOB" is a date and "SSN"
    # must not be treated as a number.
    kinds = {header: (column_kinds or {}).get(header, "GENERAL") for header in headers}
    typed_headers = [h for h in headers if kinds[h] != "GENERAL"]

    row_count = 0
    for row_data in rows:
        row_data.pop("__warnings__", None)
        if typed_headers:
            sheet.append(
                [
                    _typed_cell(sheet, row_data.get(header, ""), kinds[header])
                    for header in headers
                ]
            )
        else:
            sheet.append([sanitize_cell(row_data.get(header, "")) for header in headers])
        row_count += 1
        if row_count >= MAX_SHEET_ROWS - 1:
            raise ValueError(
                "This file exceeds the {limit:,} row Excel sheet limit. "
                "Split the conversion or export CSV instead.".format(limit=MAX_SHEET_ROWS)
            )

    workbook.save(absolute_path)
    workbook.close()

    size_bytes = os.path.getsize(absolute_path)
    relative_path = os.path.relpath(absolute_path, str(settings.MEDIA_ROOT))
    logger.info(
        "wrote workbook %s rows=%d bytes=%d", os.path.basename(absolute_path), row_count, size_bytes
    )

    return GeneratedWorkbook(
        absolute_path=absolute_path,
        relative_path=relative_path.replace(os.sep, "/"),
        filename=os.path.basename(absolute_path),
        row_count=row_count,
        size_bytes=size_bytes,
    )
