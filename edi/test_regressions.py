"""
Regression tests for the twenty defects in the work order.

Each test is named for the bug it locks down and fails against the code as it
was. They run on a synthetic interchange rather than a client file, because a
fixture built from real 834 data is PHI and must not live in a repository.

The numbering in the docstrings matches the Required Tests list in the brief.
"""

import os
import json
import tempfile

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from openpyxl import load_workbook

from conversion.models import ConversionHistory
from edi.services.element_codes import element_position, normalize_element
from edi.services.loop_extractor import extract_loops
from edi.services.parser import EDI834Parser
from edi.services.row_builder import build_excel_rows
from files.models import VALIDATED_STATUSES, GeneratedFile, ProcessingStatus, UploadedFile
from mapping.models import MappingTemplate
from members.models import Member, MemberDailyStatus, ssn_fingerprint

ISA = (
    "ISA*00*          *00*          *ZZ*SENDER         *ZZ*RECEIVER       "
    "*260801*0839*^*00501*000000001*0*P*:~"
)
GS = "GS*BE*SENDER*RECEIVER*20260801*0839*1*X*005010X220A1~"


def build_834(loops, control="0001", file_date="20260801", set_id="834"):
    """
    Assemble a structurally valid interchange.

    SE01 counts ST through SE inclusive and the validator checks it, so the
    count is computed rather than guessed — a hand-written fixture with a wrong
    SE01 fails validation for the wrong reason and the test proves nothing.
    """
    header = (
        "ST*{sid}*{c}*005010X220A1~"
        "BGN*00*{c}*{d}*0839****4~"
        "N1*P5*ACME TRUST*FI*123456789~"
    ).format(sid=set_id, c=control, d=file_date)

    body = header + "".join(loops)
    segment_count = len([s for s in body.split("~") if s.strip()]) + 1  # + SE

    return (
        ISA
        + GS
        + body
        + "SE*{n}*{c}~GE*1*1~IEA*1*000000001~".format(n=segment_count, c=control)
    )


# A subscriber loop carrying three REF segments and two DTP segments, which is
# the shape that broke first-match-wins resolution.
SUBSCRIBER = (
    "INS*Y*18*021*28*A***FT~"
    "REF*0F*400112233~"
    "REF*1L*GRP-4410~"
    "REF*17*MEM001~"
    "NM1*IL*1*WHITFIELD*MARCUS*A***34*111223333~"
    "N3*1420 FAIRGROUND RD~"
    "N4*SPRINGBORO*OH*45066~"
    "DMG*D8*19790412*M~"
    "HD*021**HLT*PPO-STD~"
    "DTP*348*D8*20260101~"
    "DTP*349*D8*20261231~"
)

DEPENDENT = (
    "INS*N*01*021*28*A~"
    "REF*0F*400112233~"
    "REF*1L*GRP-4410~"
    "NM1*IL*1*WHITFIELD*RENATA*E***34*444556666~"
    "DMG*D8*19810930*F~"
    "HD*021**HLT*PPO-STD~"
    "DTP*348*D8*20260201~"
)

# A dependent arriving with no subscriber loop anywhere in the file.
ORPHAN_DEPENDENT = (
    "INS*N*19*021*28*A~"
    "REF*0F*400112233~"
    "NM1*IL*1*WHITFIELD*IRIS****34*777889999~"
    "DMG*D8*20060214*F~"
    "HD*021**HLT*PPO-STD~"
    "DTP*348*D8*20260301~"
)

VALID_834 = build_834([SUBSCRIBER, DEPENDENT])
# Same envelope, wrong transaction set: an 835 remittance advice.
INVALID_835 = build_834([SUBSCRIBER], set_id="835")


def write_temp(content, suffix=".x12"):
    handle = tempfile.NamedTemporaryFile(
        "w", suffix=suffix, delete=False, encoding="utf-8"
    )
    handle.write(content)
    handle.close()
    return handle.name


def loops_of(content):
    path = write_temp(content)
    try:
        return extract_loops(EDI834Parser(path).iter_segments())
    finally:
        os.unlink(path)


def rule(column, segment, element, **extra):
    payload = {"excel_column": column, "segment": segment, "element": element}
    payload.update(extra)
    return payload


class ElementCodeFormatTests(TestCase):
    """Tests 1 and 2 — the hyphenated and canonical spellings must agree."""

    def test_01_hyphenated_element_code_is_normalised(self):
        """
        NM1-03 used to reach the resolver intact. It stripped 'NM1' off the
        front and parsed '-03' as an integer, which failed, so the rule
        returned an empty string for every row without raising anything.
        """
        self.assertEqual(normalize_element("NM1-03", "NM1"), "NM103")
        self.assertEqual(normalize_element("NM1-04", "NM1"), "NM104")
        self.assertEqual(normalize_element("REF-02", "REF"), "REF02")
        self.assertEqual(normalize_element("DMG-02", "DMG"), "DMG02")
        self.assertEqual(normalize_element("DTP-03", "DTP"), "DTP03")
        self.assertEqual(element_position("NM1-03", "NM1"), 3)

    def test_02_canonical_element_code_still_works(self):
        self.assertEqual(normalize_element("NM103", "NM1"), "NM103")
        self.assertEqual(element_position("NM103", "NM1"), 3)
        self.assertEqual(element_position("DMG03", "DMG"), 3)

    def test_01b_both_spellings_resolve_to_the_same_value(self):
        parsed = loops_of(VALID_834)
        hyphenated = build_excel_rows(
            parsed,
            [rule("LAST", "NM1", "NM1-03", qualifier_element="NM1-01", qualifier_value="IL")],
        )
        canonical = build_excel_rows(
            parsed,
            [rule("LAST", "NM1", "NM103", qualifier_element="NM101", qualifier_value="IL")],
        )
        self.assertEqual(hyphenated[0]["LAST"], "WHITFIELD")
        self.assertEqual(hyphenated, canonical)

    def test_different_elements_produce_different_values(self):
        """NM103 and NM104 must not be interchangeable; that is the whole point."""
        parsed = loops_of(VALID_834)
        last = build_excel_rows(parsed, [rule("NAME", "NM1", "NM1-03")])[0]["NAME"]
        first = build_excel_rows(parsed, [rule("NAME", "NM1", "NM1-04")])[0]["NAME"]
        self.assertEqual(last, "WHITFIELD")
        self.assertEqual(first, "MARCUS")
        self.assertNotEqual(last, first)


class DMGDefinitionTests(TestCase):
    """Tests 3 and 4 — DMG*D8*19790412*M~ means format, birth date, gender."""

    def test_03_dob_comes_from_dmg02(self):
        rows = build_excel_rows(loops_of(VALID_834), [rule("DOB", "DMG", "DMG-02")])
        self.assertEqual(rows[0]["DOB"], "19790412")

    def test_04_gender_comes_from_dmg03(self):
        rows = build_excel_rows(loops_of(VALID_834), [rule("SEX", "DMG", "DMG-03")])
        self.assertEqual(rows[0]["SEX"], "M")

    def test_dmg01_is_the_format_qualifier_not_the_gender(self):
        """The UI had DMG01 as gender, which would have emitted 'D8'."""
        rows = build_excel_rows(loops_of(VALID_834), [rule("FMT", "DMG", "DMG-01")])
        self.assertEqual(rows[0]["FMT"], "D8")


class QualifierResolutionTests(TestCase):
    """Tests 5 to 8 — repeated segments need the qualifier to disambiguate."""

    def setUp(self):
        self.parsed = loops_of(VALID_834)

    def test_05_ref01_0f_selects_the_subscriber_number(self):
        rows = build_excel_rows(
            self.parsed,
            [rule("SSN", "REF", "REF-02", qualifier_element="REF-01", qualifier_value="0F")],
        )
        self.assertEqual(rows[0]["SSN"], "400112233")

    def test_06_dtp01_348_selects_the_effective_date(self):
        rows = build_excel_rows(
            self.parsed,
            [rule("EFF", "DTP", "DTP-03", qualifier_element="DTP-01", qualifier_value="348")],
        )
        self.assertEqual(rows[0]["EFF"], "20260101")

    def test_07_dtp01_349_selects_the_termination_date(self):
        rows = build_excel_rows(
            self.parsed,
            [rule("TERM", "DTP", "DTP-03", qualifier_element="DTP-01", qualifier_value="349")],
        )
        self.assertEqual(rows[0]["TERM"], "20261231")

    def test_08_three_ref_segments_yield_three_distinct_values(self):
        """
        The browser parser used loopSegs.find(), taking the first REF whatever
        the column asked for, so SSN, group and member id all returned the
        REF*0F value.
        """
        rows = build_excel_rows(
            self.parsed,
            [
                rule("SSN", "REF", "REF-02", qualifier_element="REF-01", qualifier_value="0F"),
                rule("GROUP", "REF", "REF-02", qualifier_element="REF-01", qualifier_value="1L"),
                rule("MEMBER ID", "REF", "REF-02", qualifier_element="REF-01", qualifier_value="17"),
            ],
        )
        row = rows[0]
        self.assertEqual(row["SSN"], "400112233")
        self.assertEqual(row["GROUP"], "GRP-4410")
        self.assertEqual(row["MEMBER ID"], "MEM001")
        self.assertEqual(len({row["SSN"], row["GROUP"], row["MEMBER ID"]}), 3)

    def test_08b_both_dtp_dates_differ_within_one_row(self):
        rows = build_excel_rows(
            self.parsed,
            [
                rule("EFF", "DTP", "DTP-03", qualifier_element="DTP-01", qualifier_value="348"),
                rule("TERM", "DTP", "DTP-03", qualifier_element="DTP-01", qualifier_value="349"),
            ],
        )
        self.assertNotEqual(rows[0]["EFF"], rows[0]["TERM"])


def upload_and_validate(client, name, content):
    """Upload then validate through the API, as the portal now does in two steps."""
    created = client.post(
        "/api/edi/upload/",
        {"file": SimpleUploadedFile(name, content.encode())},
    ).json()
    if created.get("uploaded_file_id"):
        client.post(
            "/api/edi/validate/",
            json.dumps({"uploaded_file_id": created["uploaded_file_id"]}),
            content_type="application/json",
        )
    return created


class ApiRegressionTests(TestCase):
    """Tests 9 to 19 — the endpoints, under a real session."""

    def setUp(self):
        self.user = get_user_model().objects.create_user("owner", password="pw")
        self.client.force_login(self.user)

    def upload(self, content=VALID_834, name="t.x12"):
        """Upload then validate, matching the portal's two-step flow."""
        created = self.client.post(
            "/api/edi/upload/",
            {"file": SimpleUploadedFile(name, content.encode(), content_type="text/plain")},
        )
        body = created.json()
        if created.status_code == 201 and body.get("uploaded_file_id"):
            validated = self.client.post(
                "/api/edi/validate/",
                json.dumps({"uploaded_file_id": body["uploaded_file_id"]}),
                content_type="application/json",
            )
            merged = dict(body)
            vbody = validated.json()
            merged["status"] = vbody.get("status", merged.get("status"))
            merged["is_valid"] = vbody.get("is_valid")
            merged["errors"] = vbody.get("errors", [])
            merged["warnings"] = vbody.get("warnings", [])
            merged["member_loop_count"] = vbody.get("member_loop_count")
            merged["members_synced"] = vbody.get("members_synced", 0)
            created.json = lambda m=merged: m
        return created

    def convert(self, file_id, mappings, columns=None):
        payload = {"uploaded_file_id": file_id, "mappings": mappings}
        if columns is not None:
            payload["columns"] = columns
        return self.client.post(
            "/api/edi/convert/", payload, content_type="application/json"
        )

    # ---------------------------------------------------------------

    def test_09_an_835_cannot_pass_834_validation(self):
        with tempfile.TemporaryDirectory() as media:
            with override_settings(MEDIA_ROOT=media):
                response = self.upload(INVALID_835, "remit.x12")
                body = response.json()
                self.assertFalse(body["is_valid"])
                self.assertEqual(body["status"], ProcessingStatus.QUARANTINED)
                self.assertTrue(
                    any("834" in err for err in body["errors"]),
                    body["errors"],
                )

    def test_10_a_quarantined_file_creates_no_member_records(self):
        """
        The old upload view called sync_uploaded_file() straight after setting
        QUARANTINED, so a rejected file still wrote eligibility history.
        """
        with tempfile.TemporaryDirectory() as media:
            with override_settings(MEDIA_ROOT=media):
                self.upload(INVALID_835, "remit.x12")
                self.assertEqual(Member.objects.count(), 0)
                self.assertEqual(MemberDailyStatus.objects.count(), 0)

    def test_10b_a_valid_file_does_create_member_records(self):
        """The counterpart, so the gate above is not passing by doing nothing."""
        with tempfile.TemporaryDirectory() as media:
            with override_settings(MEDIA_ROOT=media):
                self.upload()
                self.assertEqual(Member.objects.count(), 2)
                self.assertEqual(MemberDailyStatus.objects.count(), 2)

    def test_11_a_quarantined_file_cannot_be_converted(self):
        with tempfile.TemporaryDirectory() as media:
            with override_settings(MEDIA_ROOT=media):
                created = self.upload(INVALID_835, "remit.x12").json()
                response = self.convert(
                    created["uploaded_file_id"],
                    [rule("LAST", "NM1", "NM1-03")],
                )
                self.assertEqual(response.status_code, 422, response.content)
                self.assertEqual(GeneratedFile.objects.count(), 0)

    def test_12_generated_workbook_contains_the_mapped_values(self):
        """
        End to end: the file the user downloads carries the values the mapping
        selected, resolved with qualifiers, in the requested columns.
        """
        with tempfile.TemporaryDirectory() as media:
            with override_settings(MEDIA_ROOT=media):
                created = self.upload().json()
                response = self.convert(
                    created["uploaded_file_id"],
                    [
                        rule("LAST NAME", "NM1", "NM1-03",
                             qualifier_element="NM1-01", qualifier_value="IL"),
                        rule("FIRST NAME", "NM1", "NM1-04",
                             qualifier_element="NM1-01", qualifier_value="IL"),
                        rule("DOB", "DMG", "DMG-02"),
                        rule("SEX", "DMG", "DMG-03"),
                        rule("SSN", "REF", "REF-02",
                             qualifier_element="REF-01", qualifier_value="0F"),
                        rule("EFF DATE", "DTP", "DTP-03",
                             qualifier_element="DTP-01", qualifier_value="348"),
                    ],
                )
                self.assertEqual(response.status_code, 200, response.content)
                generated = GeneratedFile.objects.get(pk=response.json()["generated_file_id"])

                workbook = load_workbook(generated.stored_file.path, read_only=True)
                sheet = workbook.worksheets[0]
                rows = list(sheet.iter_rows(values_only=True))
                workbook.close()

                headers = list(rows[0])
                subscriber = dict(zip(headers, rows[1]))

                self.assertEqual(subscriber["LAST NAME"], "WHITFIELD")
                self.assertEqual(subscriber["FIRST NAME"], "MARCUS")
                self.assertEqual(subscriber["DOB"], "19790412")
                self.assertEqual(subscriber["SEX"], "M")
                self.assertEqual(subscriber["SSN"], "400112233")
                self.assertEqual(subscriber["EFF DATE"], "20260101")

    def test_12b_changing_the_mapping_changes_the_workbook(self):
        """
        The acceptance criterion from the brief: First Name = NM104 and
        First Name = NM103 must produce different cells.
        """
        with tempfile.TemporaryDirectory() as media:
            with override_settings(MEDIA_ROOT=media):
                created = self.upload().json()
                file_id = created["uploaded_file_id"]

                def first_cell(element):
                    response = self.convert(
                        file_id,
                        [rule("FIRST NAME", "NM1", element,
                              qualifier_element="NM1-01", qualifier_value="IL")],
                    )
                    self.assertEqual(response.status_code, 200, response.content)
                    generated = GeneratedFile.objects.get(
                        pk=response.json()["generated_file_id"]
                    )
                    workbook = load_workbook(generated.stored_file.path, read_only=True)
                    value = list(workbook.worksheets[0].iter_rows(values_only=True))[1][0]
                    workbook.close()
                    return value

                self.assertEqual(first_cell("NM1-04"), "MARCUS")
                self.assertEqual(first_cell("NM1-03"), "WHITFIELD")

    def test_13_unmapped_static_headers_remain_with_blank_cells(self):
        """
        LOCAL and CLASS are deliberately unmapped. The request builder filtered
        on segment and element, so they vanished from the workbook entirely
        instead of appearing empty.
        """
        with tempfile.TemporaryDirectory() as media:
            with override_settings(MEDIA_ROOT=media):
                created = self.upload().json()
                response = self.convert(
                    created["uploaded_file_id"],
                    [rule("LAST NAME", "NM1", "NM1-03")],
                    columns=["LAST NAME", "LOCAL", "CLASS"],
                )
                self.assertEqual(response.status_code, 200, response.content)
                generated = GeneratedFile.objects.get(
                    pk=response.json()["generated_file_id"]
                )

                workbook = load_workbook(generated.stored_file.path, read_only=True)
                rows = list(workbook.worksheets[0].iter_rows(values_only=True))
                workbook.close()

                self.assertEqual(list(rows[0]), ["LAST NAME", "LOCAL", "CLASS"])
                body = dict(zip(rows[0], rows[1]))
                self.assertEqual(body["LAST NAME"], "WHITFIELD")
                self.assertIn(body["LOCAL"], ("", None))
                self.assertIn(body["CLASS"], ("", None))

    def test_14_user_a_cannot_search_user_bs_member(self):
        """
        Member search ran Member.objects.filter(ssn_fingerprint=...) with no
        owner in the filter, so any authenticated user who knew a nine digit
        number got the matching person's full record.
        """
        with tempfile.TemporaryDirectory() as media:
            with override_settings(MEDIA_ROOT=media):
                self.upload()
                mine = Member.objects.filter(owner=self.user).first()
                self.assertIsNotNone(mine)

                get_user_model().objects.create_user("intruder", password="pw")
                self.client.logout()
                self.client.login(username="intruder", password="pw")

                response = self.client.get(
                    "/api/members/search/", {"q": "111223333"}
                )
                self.assertEqual(response.status_code, 404, response.content)

                detail = self.client.get("/api/members/{pk}/".format(pk=mine.pk))
                self.assertEqual(detail.status_code, 404)

    def test_15_user_a_cannot_download_user_bs_generated_file(self):
        with tempfile.TemporaryDirectory() as media:
            with override_settings(MEDIA_ROOT=media):
                created = self.upload().json()
                converted = self.convert(
                    created["uploaded_file_id"], [rule("LAST", "NM1", "NM1-03")]
                ).json()

                get_user_model().objects.create_user("intruder", password="pw")
                self.client.logout()
                self.client.login(username="intruder", password="pw")

                self.assertEqual(self.client.get(converted["download_url"]).status_code, 404)
                self.assertEqual(self.client.get(converted["preview_url"]).status_code, 404)
                self.assertEqual(
                    self.client.get(
                        "/api/edi/uploads/{pk}/content/".format(
                            pk=created["uploaded_file_id"]
                        )
                    ).status_code,
                    404,
                )

    def test_16_same_member_in_two_same_date_files_keeps_both_occurrences(self):
        """
        MemberDailyStatus was unique on (member, status_date) and written with
        update_or_create, so the second file of the day overwrote the first
        file's row and the earlier appearance was unrecoverable.
        """
        with tempfile.TemporaryDirectory() as media:
            with override_settings(MEDIA_ROOT=media):
                first = build_834([SUBSCRIBER], control="0001", file_date="20260825")
                second = build_834([SUBSCRIBER], control="0002", file_date="20260825")

                self.upload(first, "morning.x12")
                self.upload(second, "afternoon.x12")

                member = Member.objects.get(ssn_fingerprint=ssn_fingerprint("111223333"))
                statuses = MemberDailyStatus.objects.filter(member=member)

                self.assertEqual(statuses.count(), 2)
                self.assertEqual(len({s.status_date for s in statuses}), 1)
                self.assertEqual(len({s.uploaded_file_id for s in statuses}), 2)

                search = self.client.get("/api/members/search/", {"q": "111223333"})
                self.assertEqual(search.status_code, 200, search.content)
                body = search.json()[0]
                self.assertEqual(body["file_count"], 2)
                self.assertEqual(
                    sorted(entry["file_name"] for entry in body["source_files"]),
                    ["afternoon.x12", "morning.x12"],
                )

    def test_17_archive_reports_the_actual_generated_file(self):
        """
        The Archive download button called parseEdi834ToRows(null, filename),
        which took the fallback path and built a workbook of invented members.
        The API must hand the UI a real generated file to fetch instead.
        """
        with tempfile.TemporaryDirectory() as media:
            with override_settings(MEDIA_ROOT=media):
                created = self.upload().json()
                converted = self.convert(
                    created["uploaded_file_id"], [rule("LAST", "NM1", "NM1-03")]
                ).json()

                history = self.client.get("/api/edi/history/").json()
                self.assertEqual(len(history), 1)
                entry = history[0]

                self.assertEqual(entry["generated_file_id"], converted["generated_file_id"])
                self.assertEqual(entry["download_url"], converted["download_url"])
                self.assertIsNotNone(entry["generated_file"])

                # And that URL really serves the stored workbook.
                download = self.client.get(entry["download_url"])
                self.assertEqual(download.status_code, 200)
                self.assertEqual(
                    download["Content-Type"],
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

                preview = self.client.get(entry["preview_url"]).json()
                self.assertEqual(preview["headers"], ["LAST"])
                self.assertEqual(preview["rows"][0]["LAST"], "WHITFIELD")

    def test_18_a_validation_failure_is_reported_as_a_failure(self):
        """
        The backend half of the fail-closed contract: an invalid file must come
        back is_valid=false and leave the stored status non-convertible, so the
        client has something unambiguous to render. The browser half — never
        turning a transport error into VALIDATED — lives in
        frontend/src/services/api.js and is covered by the Node contract test.
        """
        with tempfile.TemporaryDirectory() as media:
            with override_settings(MEDIA_ROOT=media):
                created = self.upload(INVALID_835, "remit.x12").json()
                response = self.client.post(
                    "/api/edi/validate/",
                    {"uploaded_file_id": created["uploaded_file_id"]},
                    content_type="application/json",
                )
                body = response.json()
                self.assertFalse(body["is_valid"])
                self.assertEqual(body["status"], ProcessingStatus.QUARANTINED)

                record = UploadedFile.objects.get(pk=created["uploaded_file_id"])
                self.assertNotIn(record.processing_status, VALIDATED_STATUSES)

    def test_19_a_failed_upload_can_be_retried(self):
        """
        Duplicate detection keyed on (owner, checksum) and rejected the second
        attempt even when the first had ended in FAILED, so a file that died in
        processing could never be reprocessed — every retry bounced off the
        broken record.
        """
        with tempfile.TemporaryDirectory() as media:
            with override_settings(MEDIA_ROOT=media):
                created = self.upload().json()
                record = UploadedFile.objects.get(pk=created["uploaded_file_id"])

                # Simulate the first attempt having died mid-processing.
                record.processing_status = ProcessingStatus.FAILED
                record.error_message = "boom"
                record.save(update_fields=["processing_status", "error_message"])

                retried = self.upload().json()
                self.assertFalse(retried["duplicate"])
                self.assertTrue(retried["retried"])
                self.assertEqual(retried["uploaded_file_id"], record.pk)
                self.assertEqual(retried["status"], ProcessingStatus.VALIDATED)

                # One file, one record: the retry must not fork the audit trail.
                self.assertEqual(UploadedFile.objects.count(), 1)

    def test_19b_a_successfully_processed_file_is_still_a_duplicate(self):
        with tempfile.TemporaryDirectory() as media:
            with override_settings(MEDIA_ROOT=media):
                self.upload()
                again = self.upload().json()
                self.assertTrue(again["duplicate"])
                self.assertEqual(UploadedFile.objects.count(), 1)

    def test_uploads_endpoint_restores_state_after_a_refresh(self):
        """Issue 17: the file list existed only in React state."""
        with tempfile.TemporaryDirectory() as media:
            with override_settings(MEDIA_ROOT=media):
                created = self.upload().json()
                self.convert(created["uploaded_file_id"], [rule("LAST", "NM1", "NM1-03")])

                listing = self.client.get("/api/edi/uploads/").json()
                self.assertEqual(len(listing), 1)
                entry = listing[0]
                # Part 1: CONVERTED is a stored fact, not a React variable. A
                # refresh at this point used to show the file back at its
                # pre-conversion status with no download link, for a workbook
                # that was already on disk.
                self.assertEqual(entry["status"], ProcessingStatus.CONVERTED)
                self.assertTrue(entry["is_converted"])
                self.assertIsNotNone(entry["converted_at"])
                self.assertEqual(entry["fileName"], "t.x12")
                self.assertIsNotNone(entry["download_url"])
                self.assertIsNotNone(entry["generated_file_id"])

    def test_edi_source_endpoint_returns_the_real_file(self):
        """Issue 18: the viewer fell back to a hard-coded sample interchange."""
        with tempfile.TemporaryDirectory() as media:
            with override_settings(MEDIA_ROOT=media):
                created = self.upload().json()
                response = self.client.get(
                    "/api/edi/uploads/{pk}/content/".format(
                        pk=created["uploaded_file_id"]
                    )
                )
                self.assertEqual(response.status_code, 200)
                body = response.json()
                self.assertIn("WHITFIELD*MARCUS", body["content"])
                self.assertNotIn("SPRINGBORO*OH*45066~DMG*D8*19810930", body["content"][:200])

    def test_convert_rejects_a_file_belonging_to_another_user(self):
        """Issue 10, ownership half: never trust the id in the request body."""
        with tempfile.TemporaryDirectory() as media:
            with override_settings(MEDIA_ROOT=media):
                created = self.upload().json()

                get_user_model().objects.create_user("intruder", password="pw")
                self.client.logout()
                self.client.login(username="intruder", password="pw")

                response = self.convert(
                    created["uploaded_file_id"], [rule("LAST", "NM1", "NM1-03")]
                )
                self.assertEqual(response.status_code, 404)


class MappingVersionTests(TestCase):
    """Issue 16 — a version used by a conversion is never edited again."""

    def setUp(self):
        self.user = get_user_model().objects.create_user("owner", password="pw")
        self.client.force_login(self.user)

    def test_editing_a_used_mapping_creates_a_new_version(self):
        with tempfile.TemporaryDirectory() as media:
            with override_settings(MEDIA_ROOT=media):
                self.client.post(
                    "/api/edi/mappings/",
                    [rule("FIRST NAME", "NM1", "NM1-04")],
                    content_type="application/json",
                )

                upload = upload_and_validate(self.client, "t.x12", VALID_834)

                converted = self.client.post(
                    "/api/edi/convert/",
                    {"uploaded_file_id": upload["uploaded_file_id"]},
                    content_type="application/json",
                ).json()
                self.assertEqual(converted["mapping_version"], 1)

                # Now change the rule that version 1 used.
                self.client.post(
                    "/api/edi/mappings/",
                    [rule("FIRST NAME", "NM1", "NM1-03")],
                    content_type="application/json",
                )

                templates = MappingTemplate.objects.filter(
                    owner=self.user, mapping_name="Default"
                ).order_by("version")
                self.assertEqual([t.version for t in templates], [1, 2])

                v1 = templates[0]
                self.assertIsNotNone(v1.locked_at)
                self.assertEqual(
                    v1.details.get(excel_column="FIRST NAME").element, "NM104"
                )
                self.assertEqual(
                    templates[1].details.get(excel_column="FIRST NAME").element, "NM103"
                )

                # History still describes what actually ran.
                record = ConversionHistory.objects.get(pk=converted["conversion_id"])
                self.assertEqual(record.mapping_template_id, v1.pk)
                self.assertEqual(record.mapping_version, 1)


class OrphanDependentTests(TestCase):
    """Issue 14 — a dependent must never be relabelled as a subscriber."""

    def setUp(self):
        self.user = get_user_model().objects.create_user("owner", password="pw")
        self.client.force_login(self.user)

    def test_dependent_arriving_before_its_subscriber_stays_a_dependent(self):
        with tempfile.TemporaryDirectory() as media:
            with override_settings(MEDIA_ROOT=media):
                orphan_file = build_834([ORPHAN_DEPENDENT], control="0001")
                upload_and_validate(self.client, "orphan.x12", orphan_file)

                member = Member.objects.get(ssn_fingerprint=ssn_fingerprint("777889999"))
                self.assertEqual(member.member_type, "DEP")
                self.assertIsNone(member.subscriber_id)
                self.assertTrue(member.subscriber_pending)

    def test_the_subscriber_arriving_later_relinks_the_dependent(self):
        with tempfile.TemporaryDirectory() as media:
            with override_settings(MEDIA_ROOT=media):
                upload_and_validate(
                    self.client, "orphan.x12", build_834([ORPHAN_DEPENDENT], control="0001")
                )
                upload_and_validate(
                    self.client, "family.x12", build_834([SUBSCRIBER], control="0002")
                )

                dependent = Member.objects.get(ssn_fingerprint=ssn_fingerprint("777889999"))
                subscriber = Member.objects.get(ssn_fingerprint=ssn_fingerprint("111223333"))

                self.assertEqual(dependent.member_type, "DEP")
                self.assertEqual(dependent.subscriber_id, subscriber.pk)
                self.assertFalse(dependent.subscriber_pending)
                self.assertEqual(subscriber.member_type, "SUB")

    def test_dependent_after_subscriber_is_linked_in_one_file(self):
        with tempfile.TemporaryDirectory() as media:
            with override_settings(MEDIA_ROOT=media):
                upload_and_validate(self.client, "t.x12", VALID_834)
                dependent = Member.objects.get(ssn_fingerprint=ssn_fingerprint("444556666"))
                subscriber = Member.objects.get(ssn_fingerprint=ssn_fingerprint("111223333"))
                self.assertEqual(dependent.member_type, "DEP")
                self.assertEqual(dependent.subscriber_id, subscriber.pk)


class MemberHistoryIndependenceTests(TestCase):
    """
    Member history comes from the X12, not from the Excel mapping.

    Stated in the brief as an architectural requirement, and worth a test
    because the two used to share a request path.
    """

    def setUp(self):
        self.user = get_user_model().objects.create_user("owner", password="pw")
        self.client.force_login(self.user)

    def test_changing_the_mapping_does_not_change_stored_member_data(self):
        with tempfile.TemporaryDirectory() as media:
            with override_settings(MEDIA_ROOT=media):
                created = self.client.post(
                    "/api/edi/upload/",
                    {"file": SimpleUploadedFile("t.x12", VALID_834.encode())},
                ).json()

                before = {
                    (m.ssn_fingerprint, m.first_name, m.last_name, m.date_of_birth, m.gender_code)
                    for m in Member.objects.all()
                }

                for element in ("NM1-03", "NM1-04", "NM1-05"):
                    self.client.post(
                        "/api/edi/convert/",
                        {
                            "uploaded_file_id": created["uploaded_file_id"],
                            "mappings": [rule("NAME", "NM1", element)],
                        },
                        content_type="application/json",
                    )

                after = {
                    (m.ssn_fingerprint, m.first_name, m.last_name, m.date_of_birth, m.gender_code)
                    for m in Member.objects.all()
                }
                self.assertEqual(before, after)
